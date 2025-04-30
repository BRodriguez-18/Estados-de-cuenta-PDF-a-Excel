import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

MESES = {"ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"}
NUM_CHAR = re.compile(r"^[0-9,().-]$")

def es_fecha(tok: str) -> bool:
    return bool(re.match(r"^\d{1,2}-[A-Z]{3}-\d{4}$", tok)) and tok[3:6] in MESES

def es_monto(txt: str) -> bool: #debe funcionar con ambos formatos
    # return bool(re.match(r"^[\d,().-]+\.\d{2}$", txt.strip()))
    t = txt.strip()
    # debe terminar en .dd
    if not re.match(r"^[\d.,()\-]+?\.\d{2}$", t):
        return False
    # al menos un dígito antes del decimal
    return True

def monto_float(txt: str) -> float: #solo se cambio t por txt
    txt = txt.strip() #y se agrego está línea
    sign = -1 if (txt.startswith("(") and txt.endswith(")")) or txt.startswith("-") else 1
    txt = txt.strip("()-").replace(",", "")
    return sign * float(txt)
   


def monto_float_2(txt: str) -> float:
    # txt=txt.strip().replace(',', '')
    # sign=-1 if txt.startswith('(') or txt.startswith('-') else 1
    # txt=txt.strip('()-')
    # return sign*float(txt)
    t = txt.strip()
    sign = -1 if (t.startswith("(") and t.endswith(")")) or t.startswith("-") else 1
    # quita paréntesis o guión
    t = t.strip("()-")
    # separar por el último punto
    if "." in t:
        int_part, dec_part = t.rsplit(".", 1)
    else:
        # sin decimal explícito
        int_part, dec_part = t, ""
    # eliminar separadores de miles
    int_part = int_part.replace(".", "").replace(",", "")
    normalized = int_part + ("." + dec_part if dec_part else "")
    return sign * float(normalized) 
def dist(a: float, b: float) -> float:
    return abs(a - b)

def unir_tokens_numericos(words, umbral: float = 1.5):
    from utils_santander import _combina
    fusion, buf = [], []
    for w in sorted(words, key=lambda w: w['x0']):
        if NUM_CHAR.match(w['text']):
            if buf and (w['x0'] - buf[-1]['x1']) > umbral:
                fusion.append(_combina(buf)); buf = []
            buf.append(w)
        else:
            if buf:
                fusion.append(_combina(buf)); buf = []
            fusion.append(w)
    if buf:
        fusion.append(_combina(buf))
    return fusion

def _combina(frags):
    texto = "".join(f['text'] for f in frags)
    nuevo = frags[0].copy()
    nuevo.update({
        'text': texto,
        'x0': frags[0]['x0'],
        'x1': frags[-1]['x1'],
        'top': min(f['top'] for f in frags),
        'bottom': max(f['bottom'] for f in frags)
    })
    return nuevo

def _formatear_excel(xls_path: str, empresa: str, no_cli: str, periodo: str, rfc: str):
    wb = load_workbook(xls_path)
    ws = wb.active
    ws.insert_rows(1, 6)
    ws["A1"] = "Banco: Santander"
    ws["A2"] = f"Empresa: {empresa}"
    ws["A3"] = f"No. Cliente: {no_cli}"
    ws["A4"] = f"Periodo: {periodo}"
    ws["A5"] = f"RFC: {rfc}"
    thin = Side("thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill("solid", "000080")
    white_font = Font(color="FFFFFF", bold=True)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(7, c)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for r in range(8, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border
    for col_cells in ws.columns:
        width = max(len(str(cell.value)) for cell in col_cells if cell.value) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(xls_path)
