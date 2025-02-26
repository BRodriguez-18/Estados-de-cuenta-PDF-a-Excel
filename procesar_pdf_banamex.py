import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def es_linea_movimiento(linea):
    """
    Determina si la línea inicia un 'movimiento' nuevo
    en formato 'dd mmm' en dos tokens separados:
    - tokens[0] = día (1-2 dígitos)
    - tokens[1] = mes (3 letras mayúsculas, p. ej. DIC, ENE)
    """
    tokens = linea.split()
    if len(tokens) < 2:
        return False

    # Verificar si tokens[0] es dd y tokens[1] es mmm
    if not re.match(r'^\d{1,2}$', tokens[0]):
        return False
    if not re.match(r'^[A-Z]{3}$', tokens[1]):
        return False

    return True

def es_numero_monetario(texto):
    """
    Determina si un texto es un número tipo '100,923.30'.
    Ajusta si tu PDF usa otro formato (p.ej. 100.923,30).
    """
    return bool(re.match(r'^[\d,]+\.\d{2}$', texto.strip()))

def dist(a, b):
    """Distancia absoluta entre dos valores."""
    return abs(a - b)

def cargar_archivo():
    global pdf_path
    archivo = filedialog.askopenfilename(
        title="Selecciona un archivo PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivo:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        entry_archivo.insert(0, archivo)
        entry_archivo.config(state=tk.DISABLED)
        pdf_path = archivo

def procesar_pdf():
    global pdf_path
    if not pdf_path:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                messagebox.showinfo("Info", "El PDF está vacío.")
                return

            # =======================
            # 1) DETECTAR ENCABEZADOS EN LA 1RA PÁGINA
            # =======================
            page0 = pdf.pages[0]
            words_page0 = page0.extract_words()

            encabezados_buscar = ["RETIROS", "DEPOSITOS", "SALDO"]
            MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
                "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

            col_positions = {}

            # Agrupamos las palabras de la primera página por 'top' para formar líneas
            lineas_dict_page0 = {}
            for w in words_page0:
                top_approx = int(w['top'])
                if top_approx not in lineas_dict_page0:
                    lineas_dict_page0[top_approx] = []
                lineas_dict_page0[top_approx].append(w)

            lineas_ordenadas_page0 = sorted(lineas_dict_page0.items(), key=lambda x: x[0])

            # Buscamos la línea que contenga los 3 encabezados
            for top_val, words_in_line in lineas_ordenadas_page0:
                line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
                # Si en esta línea aparecen los 3 encabezados, es la línea real de columnas
                if all(h in line_text_upper for h in encabezados_buscar):
                    # Extraemos la coordenada de cada encabezado
                    for w in words_in_line:
                        w_text_upper = w['text'].strip().upper()
                        if w_text_upper in encabezados_buscar:
                            center_x = (w['x0'] + w['x1']) / 2
                            col_positions[w_text_upper] = center_x
                    break

            # Ordenamos por la coordenada X
            columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])
            print("columnas_ordenadas =", columnas_ordenadas)

            # =======================
            # 2) VARIABLES PARA ENCABEZADOS DEL EXCEL
            # =======================
            periodo_str = ""
            no_cuenta_str = ""
            empresa_str = ""
            no_cliente_str = ""
            rfc_str = ""

            # =======================
            # 3) FRASES A OMITIR (skip) Y A DETENER (stop)
            # =======================
            skip_phrases = [
                "ESTADO DE CUENTA AL",
                "Página",
            ]
            skip_phrases = [s.upper() for s in skip_phrases]

            stop_phrases = [
                "SALDO MINIMO REQUERIDO",
                # "COMISIONES EFECTIVAMENTE COBRADAS"
            ]

            start_reading = False
            stop_reading = False

            todos_los_movimientos = []
            movimiento_actual = None

            # =======================
            # 4) RECORRER TODAS LAS PÁGINAS PARA DETECTAR MOVIMIENTOS
            # =======================
            for page_index, page in enumerate(pdf.pages):
                if stop_reading:
                    break

                words = page.extract_words()
                # Agrupamos por 'top'
                lineas_dict = {}
                for w in words:
                    top_approx = int(w['top'])
                    if top_approx not in lineas_dict:
                        lineas_dict[top_approx] = []
                    lineas_dict[top_approx].append(w)

                lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                for top_val, words_in_line in lineas_ordenadas:
                    if stop_reading:
                        break

                    line_text = " ".join(w['text'] for w in words_in_line)
                    line_text_upper = line_text.upper()

                    # Detectar periodo, p. ej. "RESUMEN DEL: 01/DIC/2023 AL 31/DIC/2023"
                    if "RESUMEN" in line_text_upper and "DEL:" in line_text_upper:
                        tokens_line = line_text.split()
                        fechas = [t for t in tokens_line if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', t)]
                        if len(fechas) == 2:
                            periodo_str = f"{fechas[0]} al {fechas[1]}"
                        else:
                            periodo_str = line_text
                        continue

                    # Detectar número de cuenta => "CONTRATO 12300415060"
                    if "CONTRATO" in line_text_upper and not no_cuenta_str:
                        tokens_line = line_text.split()
                        no_cuenta_str = tokens_line[-1]
                        continue

                    # Detectar cliente => "CLIENTE: 2855558"
                    if "CLIENTE:" in line_text_upper and not no_cliente_str:
                        tokens_line = line_text.split()
                        no_cliente_str = tokens_line[-1]
                        continue

                    # Detectar RFC => "Registro Federal de Contribuyentes: HST781101TJ8"
                    if "REGISTRO FEDERAL DE CONTRIBUYENTES:" in line_text_upper and not rfc_str:
                        tokens_line = line_text.split()
                        rfc_str = tokens_line[-1]
                        continue

                    # Revisar stop_phrases
                    if any(sp in line_text_upper for sp in stop_phrases):
                        stop_reading = True
                        break

                    # Empezar a leer movimientos cuando detectemos la primera fecha "dd mmm"
                    if not start_reading:
                        tokens_line = line_text.split()
                        found_day = any(re.match(r'^\d{1,2}$', t) for t in tokens_line)
                        found_month = any(re.match(r'^[A-Z]{3}$', t) for t in tokens_line)
                        if found_day and found_month:
                            start_reading = True
                        else:
                            # Aún no es movimiento, saltar
                            continue

                    # Omitir líneas con skip_phrases
                    if any(sp in line_text_upper for sp in skip_phrases):
                        continue

                    # ¿Es un nuevo movimiento? => tokens[0] = dd, tokens[1] = mmm
                    if es_linea_movimiento(line_text_upper):
                        # Guardar el anterior
                        if movimiento_actual:
                            todos_los_movimientos.append(movimiento_actual)

                        tokens_line = line_text_upper.split()
                        movimiento_actual = {
                            "Fecha": f"{tokens_line[0]} {tokens_line[1]}",
                            "Concepto": "",
                            "Retiros": None,
                            "Depositos": None,
                            "Saldo": None
                        }
                    else:
                        # Continuación
                        if not movimiento_actual:
                            movimiento_actual = {
                                "Fecha": None,
                                "Concepto": "",
                                "Retiros": None,
                                "Depositos": None,
                                "Saldo": None
                            }

                    # Asignar montos por coordenadas
                    for w in words_in_line:
                        txt = w['text'].strip()
                        center_w = (w['x0'] + w['x1']) / 2

                        if es_numero_monetario(txt):
                            # Ubicar la columna más cercana
                            if columnas_ordenadas:
                                col_name, col_center = min(
                                    columnas_ordenadas,
                                    key=lambda x: dist(x[1], center_w)
                                )
                                if col_name == "RETIROS":
                                    movimiento_actual["Retiros"] = txt
                                elif col_name == "DEPOSITOS":
                                    movimiento_actual["Depositos"] = txt
                                elif col_name == "SALDO":
                                    movimiento_actual["Saldo"] = txt
                            else:
                                # Si no detectamos columnas, asume Retiros
                                movimiento_actual["Retiros"] = txt
                        else:
                            # Texto al concepto (omitir dd y mmm)
                            if re.match(r'^\d{1,2}$', txt) or txt in MESES_CORTOS:
                                continue
                            movimiento_actual["Concepto"] += " " + txt

            # Al terminar
            if movimiento_actual:
                todos_los_movimientos.append(movimiento_actual)

        # =======================
        # 5) GUARDAR EN EXCEL
        # =======================
        df = pd.DataFrame(todos_los_movimientos, columns=[
            "Fecha",
            "Concepto",
            "Retiros",
            "Depositos",
            "Saldo"
        ])

        ruta_salida = "movimientos_citibanamex_dd_mmm.xlsx"
        df.to_excel(ruta_salida, index=False)

        # Ajustes de estilo con openpyxl
        wb = load_workbook(ruta_salida)
        ws = wb.active

        # Insertar filas para encabezado
        ws.insert_rows(1, 6)

        # Encabezado
        ws["A1"] = f"Banco: Citibanamex"
        ws["A2"] = f"Empresa: {empresa_str}"
        ws["A3"] = f"No. Cuenta: {no_cuenta_str}"
        ws["A4"] = f"No. Cliente: {no_cliente_str}"
        ws["A5"] = f"Periodo: {periodo_str}"
        ws["A6"] = f"RFC: {rfc_str}"

        thin_side = Side(border_style="thin")
        thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)

        max_row = ws.max_row
        max_col = ws.max_column

        # Estilo para la fila de encabezados (fila 7)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=7, column=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Estilo para filas de datos
        for row in range(8, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                # if col in [3, 4, 5]:
                #     cell.alignment = Alignment(horizontal="right")

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2

        # Alineación wrap_text
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        wb.save(ruta_salida)
        messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

# Interfaz gráfica con tkinter
root = tk.Tk()
root.title("Extracción Movimientos - dd mmm (Encabezados en 1ra página)")
root.geometry("600x250")

pdf_path = ""

btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
btn_cargar.pack(pady=10)

entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
entry_archivo.pack(padx=10, pady=10)

btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
btn_procesar.pack(pady=10)

root.mainloop()
