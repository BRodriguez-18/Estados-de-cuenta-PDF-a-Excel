# import sys
# import os
# import tkinter as tk
# from tkinter import filedialog, messagebox
# import pdfplumber
# import pandas as pd
# import re
# import string
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# # Conjunto de meses cortos
# MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
#                 "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

# def ajusta_fechas_en_linea(line_text):
#     """
#     Inserta un espacio si se detecta una fecha en formato dd-MES-yy
#     inmediatamente seguida de letras o dígitos, p.ej. '06-ENE-25ABC' => '06-ENE-25 ABC'.
#     Solo hace una pasada, evitando bucles infinitos.
#     """
#     pattern = r'(\d{1,2}-[A-Z]{3}-\d{2})([A-Za-z0-9])'
#     new_text, num_subs = re.subn(pattern, r'\1 \2', line_text)
#     return new_text

# def es_linea_movimiento(linea):
#     """
#     Determina si la línea inicia un 'movimiento' nuevo
#     con el primer token en formato dd-MES-yy (p.ej. '03-ENE-25').
#     """
#     tokens = linea.split()
#     if not tokens:
#         return False
    
#     primer_token = tokens[0]
#     partes = primer_token.split("-")
#     if len(partes) != 3:
#         return False
    
#     dia, mes, anio = partes
#     if not re.match(r'^\d{1,2}$', dia):
#         return False
#     if mes not in MESES_CORTOS:
#         return False
#     if not re.match(r'^\d{2}$', anio):
#         return False

#     return True

# def es_numero_monetario(texto):
#     """
#     Determina si un texto es un número tipo '100,923.30'.
#     Ajusta si tu PDF usa otro formato.
#     """
#     # pattern = r'^\d{1,3}(,\d{3})*\.\d{2}$'
#     pattern = r'^-?\d{1,3}(,\d{3})*\.\d{2}-?$'
#     return bool(re.match(pattern, texto.strip()))

# def parse_monetario(txt):
#     """
#     Convierte un texto en formato '100,923.30' a un float.
#     """
#     txt = txt.strip()
#     sign = 1
    
#     if txt.endswith("-"):
#         sign = -1
#         txt = txt[:-1].strip()
#     # Verifica si está entre paréntesis (se asume negativo)
#     elif txt.startswith("(") and txt.endswith(")"):
#         sign = -1
#         txt = txt[1:-1].strip()
#     # Verifica si empieza con signo "-"
#     elif txt.startswith("-"):
#         sign = -1
#         txt = txt[1:].strip()

#     # Elimina comas antes de convertir a float
#     txt = txt.replace(",", "")
#     return sign * float(txt)

# def dist(a, b):
#     """Distancia absoluta entre dos valores."""
#     return abs(a - b)

# def cargar_archivo():
#     global pdf_paths
#     archivos = filedialog.askopenfilenames(
#         title="Selecciona uno o más archivos PDF",
#         filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
#     )
#     if archivos:
#         entry_archivo.config(state=tk.NORMAL)
#         entry_archivo.delete(0, tk.END)
#         files_text = " ; ".join(archivos)
#         entry_archivo.insert(0, files_text)
#         entry_archivo.config(state=tk.DISABLED)
#         pdf_paths = archivos

# def agrupar_por_top_con_tolerancia(words, tolerancia=2):
#     """
#     Recibe una lista de 'words' extraídas por pdfplumber y 
#     las agrupa en diccionarios de la forma {top_agrupado: [words_en_esa_linea]}.
#     La clave 'top_agrupado' es un float o int representativo de la línea.
    
#     'tolerancia' indica cuán cerca (en unidades de 'top') deben estar las palabras
#     para considerarlas parte de la misma línea.
#     """
#     lineas_dict = {}
#     for w in words:
#         actual_top = w['top']
#         top_encontrado = None
#         for top_existente in lineas_dict.keys():
#             if abs(top_existente - actual_top) <= tolerancia:
#                 top_encontrado = top_existente
#                 break
#         if top_encontrado is not None:
#             lineas_dict[top_encontrado].append(w)
#         else:
#             lineas_dict[actual_top] = [w]
#     lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])
#     return lineas_ordenadas

# def procesar_pdf():
#     global pdf_paths
#     if not pdf_paths:
#         messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
#         return
#     for pdf_path in pdf_paths:
#         try:
#             with pdfplumber.open(pdf_path) as pdf:
#                 pdf_name = os.path.basename(pdf_path)
#                 pdf_stem, pdf_ext = os.path.splitext(pdf_name)
#                 excel_name = pdf_stem + ".xlsx"

#                 if len(pdf.pages) == 0:
#                     messagebox.showinfo("Info", "El PDF está vacío.")
#                     return

#                 page0 = pdf.pages[0]
#                 words_page0 = page0.extract_words()

#                 regionEmpresa = (50.400000000000006, 63.426000000000045, 200, 72.42600000000004)
#                 regionPeriodo = (411.70875,105.45000000000005,550,114.45000000000005)
#                 croppedEmpresa = pdf.pages[0].within_bbox(regionEmpresa)
#                 croppedPeriodo = pdf.pages[0].within_bbox(regionPeriodo)
#                 empresa_str = croppedEmpresa.extract_text() or ""
#                 periodo_str = croppedPeriodo.extract_text() or ""
#                 # Debug: Mostrar las palabras extraídas de la página 1
#                 for word in words_page0:
#                     print(f"Texto: {word['text']}, x0: {word['x0']}, x1: {word['x1']}, top: {word['top']}, bottom: {word['bottom']}")

#                 # 1) DETECTAR ENCABEZADOS EN LA 2DA PÁGINA (índice 1)
#                 page1 = pdf.pages[1]
#                 words_page1 = page1.extract_words()

#                 # Debug: Mostrar las palabras extraídas de la página 2
#                 # print("[DEBUG] Palabras extraídas de la página 2:")
#                 # for word in words_page1:
#                     # print(word)
#                     # print(f"Texto: {word['text']}, x0: {word['x0']}, x1: {word['x1']}, top: {word['top']}, bottom: {word['bottom']}")

#                 encabezados_buscar = ["DEPOSITO", "RETIRO", "SALDO"]
#                 col_positions = {}

                

#                 lineas_ordenadas_page1 = agrupar_por_top_con_tolerancia(words_page1, tolerancia=2)

#                 for top_val, words_in_line in lineas_ordenadas_page1:
#                     line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
#                     if all(h in line_text_upper for h in encabezados_buscar):
#                         for w in words_in_line:
#                             w_text_upper = w['text'].strip().upper()
#                             if w_text_upper in encabezados_buscar:
#                                 center_x = (w['x0'] + w['x1']) / 2
#                                 col_positions[w_text_upper] = center_x
#                         break

#                 columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])

#                 # 2) VARIABLES PARA ENCABEZADOS DEL EXCEL
#                 # periodo_str = ""
#                 # empresa_str = ""
#                 no_cliente_str = ""
#                 rfc_str = ""

#                 # 3) FRASES A OMITIR (skip) Y A DETENER (stop)
#                 skip_phrases = [ 
#                     "ESTADO DE CUENTA",
#                     "FECHA DE CORTE",
#                     "Línea Directa para su empresa:",
#                     "CIUDAD DE MÉXICO: (55)",
#                     "(81)8156 9640",
#                     "CIUDAD DE MÉXICO:",
#                     "Guadalajara",
#                     "Monterrey",
#                     "BANCO MERCANTIL DEL NORTE S.A. INSTITUCIÓN DE BANCA MÚLTIPLE GRUPO FINANCIERO BANORTE.",
#                     "Nuevo Leon. RFC BMN930209927",
#                     "Resto del país:",
#                     "DETALLE DE MOVIMIENTOS",
#                     "Enlace Negocios Basica",
#                     "Visita nuestra página:",
#                     "FECHA",
#                     "DESCRIPCIÓN / ESTABLECIMIENTO",
#                     "MONTO DEL DEPOSITO",
#                     "MONTO DEL RETIRO",
#                     "SALDO",
#                     "Banco Mercantil del Norte",
#                 ]
#                 stop_phrases = ["OTROS"]

#                 start_reading = False
#                 stop_reading = False
#                 todos_los_movimientos = []
#                 movimiento_actual = None

#                 # 4) RECORRER TODAS LAS PÁGINAS
#                 for page_index, page in enumerate(pdf.pages):
#                     if stop_reading:
#                         break

#                     words = page.extract_words()
#                     lineas_dict = {}
#                     for w in words:
#                         top_approx = int(w['top'])
#                         if top_approx not in lineas_dict:
#                             lineas_dict[top_approx] = []
#                         lineas_dict[top_approx].append(w)

#                     lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

#                     for top_val, words_in_line in lineas_ordenadas:
#                         if stop_reading:
#                             break

#                         line_text = " ".join(w['text'] for w in words_in_line)
#                         line_text = ajusta_fechas_en_linea(line_text)
#                         line_text = re.sub(
#                             r'(\d{1,2}-[A-Z]{3}-\d{2})([A-Za-z])',
#                             r'\1 \2',
#                             line_text
#                         )
#                         line_text_upper = line_text.upper()
                        
#                         # Si estamos en páginas posteriores, saltamos el encabezado repetido
#                         if page_index >= 1:
#                             header_keywords = ["FECHA", "MONTO DEL DEPOSITO", "MONTO DEL RETIRO", "SALDO"]
#                             if all(keyword in line_text_upper for keyword in header_keywords):
#                                 continue

#                         # if "Periodo" in line_text_upper and not periodo_str:
#                         #     tokens_line = line_text.split()
                            
#                         #     if "Periodo" in tokens_line:
#                         #         idx_periodo = tokens_line.index("Periodo")
#                         #         periodo_str = " ".join(tokens_line[idx_periodo + 1:])
#                         #     continue

#                         if "NO. DE CLIENTE:" in line_text_upper and not no_cliente_str:
#                             tokens_line = line_text.split()
#                             no_cliente_str = tokens_line[-1]
#                             continue

#                         if "RFC:" in line_text_upper and not rfc_str:
#                             tokens_line = line_text.split()
#                             rfc_str = tokens_line[-1]
#                             continue

#                         if page_index >= 1 and any(sp in line_text_upper for sp in stop_phrases):
#                             stop_reading = True
#                             break

#                         if not start_reading:
#                             if es_linea_movimiento(line_text_upper):
#                                 start_reading = True
#                             else:
#                                 continue

#                         if any(sp in line_text for sp in skip_phrases):
#                             continue    

#                         if re.search(r'\b\d+/\d+\b', line_text_upper):
#                             continue

#                         if es_linea_movimiento(line_text_upper):
#                             if movimiento_actual:
#                                 todos_los_movimientos.append(movimiento_actual)
#                             tokens_line = line_text_upper.split()
#                             movimiento_actual = {
#                                 "Fecha": tokens_line[0],
#                                 "Descripción / Establecimiento": "",
#                                 "Monto del deposito": None,
#                                 "Monto del retiro": None,
#                                 "Saldo": None
#                             }
#                         else:
#                             if not movimiento_actual:
#                                 movimiento_actual = {
#                                     "Fecha": None,
#                                     "Descripción / Establecimiento": "",
#                                     "Monto del deposito": None,
#                                     "Monto del retiro": None,
#                                     "Saldo": None
#                                 }

#                         for w in words_in_line:
#                             token_upper = w['text'].upper()
#                             if any(sp in token_upper for sp in skip_phrases):
#                                 break

#                             txt = w['text'].strip()
#                             center_w = (w['x0'] + w['x1']) / 2

#                             if es_numero_monetario(txt):
#                                 val=parse_monetario(txt)
#                                 if columnas_ordenadas:
#                                     col_name, col_center = min(
#                                         columnas_ordenadas,
#                                         key=lambda x: dist(x[1], center_w)
#                                     )
#                                     # print(f"[DEBUG] -> Token '{txt}' en columna '{col_name}'")
#                                     if "DEPOSITO" in col_name.upper():
#                                         movimiento_actual["Monto del deposito"] = val
#                                     elif "RETIRO" in col_name.upper():
#                                         movimiento_actual["Monto del retiro"] = val
#                                     elif "SALDO" in col_name.upper():
#                                         movimiento_actual["Saldo"] = val
#                                     else:
#                                         movimiento_actual["Monto del retiro"] = val
#                                 else:
#                                     movimiento_actual["Monto del retiro"] = val
#                             else:
#                                 m = re.match(r'^(\d{1,2}-[A-Z]{3}-\d{2})(.*)$', txt)
#                                 if m:
#                                     date_part = m.group(1)
#                                     rest = m.group(2)
#                                     if movimiento_actual["Fecha"] and date_part.upper() == movimiento_actual["Fecha"]:
#                                         txt = rest.strip()
#                                 clean_txt = txt.strip(string.punctuation)
#                                 if re.match(r'^\d{1,2}$', clean_txt):
#                                     continue
#                                 if clean_txt in MESES_CORTOS:
#                                     continue
#                                 if re.match(r'^\d{1,2}-[A-Z]{3}-\d{2}$', clean_txt):
#                                     continue
#                                 movimiento_actual["Descripción / Establecimiento"] += txt + " "

#                 if movimiento_actual:
#                     for sp in skip_phrases:
#                         movimiento_actual["Descripción / Establecimiento"] = re.sub(
#                             re.escape(sp), "", movimiento_actual["Descripción / Establecimiento"], flags=re.IGNORECASE
#                         )
#                     todos_los_movimientos.append(movimiento_actual)

#             df = pd.DataFrame(todos_los_movimientos, columns=[
#                 "Fecha",
#                 "Descripción / Establecimiento",
#                 "Monto del deposito",
#                 "Monto del retiro",
#                 "Saldo"
#             ])

#             ruta_salida = os.path.join(output_folder, excel_name)
#             df.to_excel(ruta_salida, index=False)

#             wb = load_workbook(ruta_salida)
#             ws = wb.active

#             ws.insert_rows(1, 6)

#             ws["A1"] = f"Banco: Banorte"
#             ws["A2"] = f"Empresa: {empresa_str}"
#             ws["A3"] = f"No. Cliente: {no_cliente_str}"
#             ws["A4"] = f"Periodo: {periodo_str}"
#             ws["A5"] = f"RFC: {rfc_str}"

#             thin_side = Side(border_style="thin")
#             thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

#             header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
#             white_font = Font(color="FFFFFF", bold=True)

#             max_row = ws.max_row
#             max_col = ws.max_column

#             for col in range(1, max_col + 1):
#                 cell = ws.cell(row=7, column=col)
#                 cell.fill = header_fill
#                 cell.font = white_font
#                 cell.alignment = Alignment(horizontal="center")
#                 cell.border = thin_border

#             for row in range(8, max_row + 1):
#                 for col in range(1, max_col + 1):
#                     cell = ws.cell(row=row, column=col)
#                     cell.border = thin_border

#             for col in ws.columns:
#                 max_length = 0
#                 col_letter = col[0].column_letter
#                 for cell in col:
#                     if cell.value is not None:
#                         length = len(str(cell.value))
#                         if length > max_length:
#                             max_length = length
#                 ws.column_dimensions[col_letter].width = max_length + 2

#             for row in ws.iter_rows():
#                 for cell in row:
#                     cell.alignment = Alignment(wrap_text=True)

#             wb.save(ruta_salida)
#             messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

#         except Exception as e:
#             messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

# def main():
#     global pdf_paths, output_folder
#     if len(sys.argv) < 2:
#         # print("Uso: python procesar_pdf_banorte.py <carpeta_salida>")
#         return
#     output_folder = sys.argv[1]
#     pdf_paths = ""
#     root = tk.Tk()
#     root.title("Extracción Movimientos - Banorte")

#     win_width = 600
#     win_height = 250
#     screen_width = root.winfo_screenwidth()
#     screen_height = root.winfo_screenheight()
#     x = (screen_width - win_width) // 2
#     y = (screen_height - win_height) // 2
#     root.geometry(f"{win_width}x{win_height}+{x}+{y}")
#     root.update()
#     root.lift()
#     root.focus_force()
#     root.attributes("-topmost", True)
#     root.after(10, lambda: root.attributes("-topmost", False))

#     btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
#     btn_cargar.pack(pady=10)

#     global entry_archivo
#     entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
#     entry_archivo.pack(padx=10, pady=10)

#     btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
#     btn_procesar.pack(pady=10)

#     root.mainloop()

# if __name__ == "__main__":
#     main()


import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
import string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Conjunto de meses cortos
MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
                "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

def ajusta_fechas_en_linea(line_text):
    """
    Inserta un espacio si se detecta una fecha en formato dd-MES-yy
    inmediatamente seguida de letras o dígitos, p.ej. '06-ENE-25ABC' => '06-ENE-25 ABC'.
    Solo hace una pasada, evitando bucles infinitos.
    """
    pattern = r'(\d{1,2}-[A-Z]{3}-\d{2})([A-Za-z0-9])'
    new_text, num_subs = re.subn(pattern, r'\1 \2', line_text)
    return new_text

def es_linea_movimiento(linea):
    """
    Determina si la línea inicia un 'movimiento' nuevo
    con el primer token en formato dd-MES-yy (p.ej. '03-ENE-25').
    """
    tokens = linea.split()
    if not tokens:
        return False
    
    primer_token = tokens[0]
    partes = primer_token.split("-")
    if len(partes) != 3:
        return False
    
    dia, mes, anio = partes
    if not re.match(r'^\d{1,2}$', dia):
        return False
    if mes not in MESES_CORTOS:
        return False
    if not re.match(r'^\d{2}$', anio):
        return False

    return True

def es_numero_monetario(texto):
    """
    Determina si un texto es un número tipo '100,923.30'.
    Ajusta si tu PDF usa otro formato.
    """
    pattern = r'^-?\d{1,3}(,\d{3})*\.\d{2}-?$'
    return bool(re.match(pattern, texto.strip()))

def parse_monetario(txt):
    """
    Convierte un texto en formato '100,923.30' a un float.
    """
    txt = txt.strip()
    sign = 1
    
    if txt.endswith("-"):
        sign = -1
        txt = txt[:-1].strip()
    # Verifica si está entre paréntesis (se asume negativo)
    elif txt.startswith("(") and txt.endswith(")"):
        sign = -1
        txt = txt[1:-1].strip()
    # Verifica si empieza con signo "-"
    elif txt.startswith("-"):
        sign = -1
        txt = txt[1:].strip()

    # Elimina comas antes de convertir a float
    txt = txt.replace(",", "")
    return sign * float(txt)

def dist(a, b):
    """Distancia absoluta entre dos valores."""
    return abs(a - b)

def agrupa(words):
    d={}
    for w in words:
        d.setdefault(int(w['top']), []).append(w)
    return sorted(d.items())

def debug_page(pdf):
    if len(pdf.pages)<2: return
    print("\n── DEBUG PAGE 2 ─────────────────────────────────────")
    for top,grp in agrupa(pdf.pages[1].extract_words()):
        print(f"[top={top:>4}] "+" ".join(w['text'] for w in grp))
        for w in grp:
            print(f"     ↳ '{w['text']}'  x0={w['x0']:.1f}  x1={w['x1']:.1f}  bottom={w['bottom']:.1f}")
    print("──────────────────────────────────────────────────────\n")

def cargar_archivo():
    global pdf_paths
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos

def agrupar_por_top_con_tolerancia(words, tolerancia=2):
    """
    Recibe una lista de 'words' extraídas por pdfplumber y 
    las agrupa en diccionarios de la forma {top_agrupado: [words_en_esa_linea]}.
    La clave 'top_agrupado' es un float o int representativo de la línea.
    
    'tolerancia' indica cuán cerca (en unidades de 'top') deben estar las palabras
    para considerarlas parte de la misma línea.
    """
    lineas_dict = {}
    for w in words:
        actual_top = w['top']
        top_encontrado = None
        for top_existente in lineas_dict.keys():
            if abs(top_existente - actual_top) <= tolerancia:
                top_encontrado = top_existente
                break
        if top_encontrado is not None:
            lineas_dict[top_encontrado].append(w)
        else:
            lineas_dict[actual_top] = [w]
    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])
    return lineas_ordenadas

def procesar_pdf():
    global pdf_paths
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return
    for pdf_path in pdf_paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pdf_name = os.path.basename(pdf_path)
                pdf_stem, pdf_ext = os.path.splitext(pdf_name)
                excel_name = pdf_stem + ".xlsx"

                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return

                # Buscar la primera página que contenga "ESTADO DE CUENTA"
                first_page_index = None
                for idx, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    if "ESTADO DE CUENTA" in page_text.upper():
                        first_page_index = idx
                        break

                if first_page_index is None:
                    messagebox.showwarning("Advertencia", "No se encontró la página con 'ESTADO DE CUENTA'.")
                    return

                first_page = pdf.pages[1]

                # Extraer datos de encabezados a partir de la primera página relevante
                words_first_page = first_page.extract_words()
                regionEmpresa = (50.400000000000006, 63.426000000000045, 200, 72.42600000000004)
                regionPeriodo = (411.70875, 105.45000000000005, 550, 114.45000000000005)
                croppedEmpresa = first_page.within_bbox(regionEmpresa)
                croppedPeriodo = first_page.within_bbox(regionPeriodo)
                empresa_str = croppedEmpresa.extract_text() or ""
                periodo_str = croppedPeriodo.extract_text() or ""

                # Debug: Mostrar las palabras extraídas de la primera página relevante
                for word in words_first_page:
                    print(f"Texto: {word['text']}, x0: {word['x0']}, x1: {word['x1']}, top: {word['top']}, bottom: {word['bottom']}")

                # 1) DETECTAR ENCABEZADOS EN LA 2DA PÁGINA (o siguiente a la relevante)
                encabezados_buscar = ["DEPOSITO", "RETIRO", "SALDO"]
                col_positions = {}

                encabezados_encontrados = False
                # Se recorren las páginas siguientes a la primera relevante
                for idx in range(first_page_index + 1, len(pdf.pages)):
                    page = pdf.pages[idx]
                    words_page = page.extract_words()
                    lineas_ordenadas = agrupar_por_top_con_tolerancia(words_page, tolerancia=2)
                    for top_val, words_in_line in lineas_ordenadas:
                        line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
                        if all(h in line_text_upper for h in encabezados_buscar):
                            for w in words_in_line:
                                w_text_upper = w['text'].strip().upper()
                                if w_text_upper in encabezados_buscar:
                                    center_x = (w['x0'] + w['x1']) / 2
                                    col_positions[w_text_upper] = center_x
                            encabezados_encontrados = True
                            break
                    if encabezados_encontrados:
                        break

                columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])

                # 2) VARIABLES PARA ENCABEZADOS DEL EXCEL
                no_cliente_str = ""
                rfc_str = ""

                # 3) FRASES A OMITIR (skip) Y A DETENER (stop)
                skip_phrases = [ 
                    "ESTADO DE CUENTA",
                    "FECHA DE CORTE",
                    "Línea Directa para su empresa:",
                    "CIUDAD DE MÉXICO: (55)",
                    "(81)8156 9640",
                    "CIUDAD DE MÉXICO:",
                    "Guadalajara",
                    "Monterrey",
                    "BANCO MERCANTIL DEL NORTE S.A. INSTITUCIÓN DE BANCA MÚLTIPLE GRUPO FINANCIERO BANORTE.",
                    "Nuevo Leon. RFC BMN930209927",
                    "Resto del país:",
                    "DETALLE DE MOVIMIENTOS",
                    "Enlace Negocios Basica",
                    "Visita nuestra página:",
                    "FECHA",
                    "DESCRIPCIÓN / ESTABLECIMIENTO",
                    "MONTO DEL DEPOSITO",
                    "MONTO DEL RETIRO",
                    "SALDO",
                    "Banco Mercantil del Norte",
                ]
                stop_phrases = ["OTROS"]

                start_reading = False
                stop_reading = False
                todos_los_movimientos = []
                movimiento_actual = None

                # 4) RECORRER TODAS LAS PÁGINAS
                for page_index, page in enumerate(pdf.pages):
                    if stop_reading:
                        break

                    words = page.extract_words()
                    lineas_dict = {}
                    for w in words:
                        top_approx = int(w['top'])
                        if top_approx not in lineas_dict:
                            lineas_dict[top_approx] = []
                        lineas_dict[top_approx].append(w)

                    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                    for top_val, words_in_line in lineas_ordenadas:
                        if stop_reading:
                            break

                        line_text = " ".join(w['text'] for w in words_in_line)
                        line_text = ajusta_fechas_en_linea(line_text)
                        line_text = re.sub(
                            r'(\d{1,2}-[A-Z]{3}-\d{2})([A-Za-z])',
                            r'\1 \2',
                            line_text
                        )
                        line_text_upper = line_text.upper()
                        
                        # Saltar encabezados repetidos en páginas posteriores a la primera relevante
                        if page_index >= first_page_index + 1:
                            header_keywords = ["FECHA", "DEPOSITO", "RETIRO", "SALDO"]
                            if all(keyword in line_text_upper for keyword in header_keywords):
                                continue

                        if "NO. DE CLIENTE:" in line_text_upper and not no_cliente_str:
                            tokens_line = line_text.split()
                            no_cliente_str = tokens_line[-1]
                            continue

                        if "RFC:" in line_text_upper and not rfc_str:
                            tokens_line = line_text.split()
                            rfc_str = tokens_line[-1]
                            continue

                        if page_index >= first_page_index + 1 and any(sp in line_text_upper for sp in stop_phrases):
                            stop_reading = True
                            break

                        if not start_reading:
                            if es_linea_movimiento(line_text_upper):
                                start_reading = True
                            else:
                                continue

                        if any(sp in line_text for sp in skip_phrases):
                            continue    

                        if re.search(r'\b\d+/\d+\b', line_text_upper):
                            continue

                        if es_linea_movimiento(line_text_upper):
                            if movimiento_actual:
                                todos_los_movimientos.append(movimiento_actual)
                            tokens_line = line_text_upper.split()
                            movimiento_actual = {
                                "Fecha": tokens_line[0],
                                "Descripción / Establecimiento": "",
                                "Monto del deposito": None,
                                "Monto del retiro": None,
                                "Saldo": None
                            }
                        else:
                            if not movimiento_actual:
                                movimiento_actual = {
                                    "Fecha": None,
                                    "Descripción / Establecimiento": "",
                                    "Monto del deposito": None,
                                    "Monto del retiro": None,
                                    "Saldo": None
                                }

                        for w in words_in_line:
                            token_upper = w['text'].upper()
                            if any(sp in token_upper for sp in skip_phrases):
                                break

                            txt = w['text'].strip()
                            center_w = (w['x0'] + w['x1']) / 2

                            if es_numero_monetario(txt):
                                val = parse_monetario(txt)
                                if columnas_ordenadas:
                                    col_name, col_center = min(
                                        columnas_ordenadas,
                                        key=lambda x: dist(x[1], center_w)
                                    )
                                    if "DEPOSITO" in col_name.upper():
                                        movimiento_actual["Monto del deposito"] = val
                                    elif "RETIRO" in col_name.upper():
                                        movimiento_actual["Monto del retiro"] = val
                                    elif "SALDO" in col_name.upper():
                                        movimiento_actual["Saldo"] = val
                                    else:
                                        movimiento_actual["Monto del retiro"] = val
                                else:
                                    movimiento_actual["Monto del retiro"] = val
                            else:
                                m = re.match(r'^(\d{1,2}-[A-Z]{3}-\d{2})(.*)$', txt)
                                if m:
                                    date_part = m.group(1)
                                    rest = m.group(2)
                                    if movimiento_actual["Fecha"] and date_part.upper() == movimiento_actual["Fecha"]:
                                        txt = rest.strip()
                                clean_txt = txt.strip(string.punctuation)
                                if re.match(r'^\d{1,2}$', clean_txt):
                                    continue
                                if clean_txt in MESES_CORTOS:
                                    continue
                                if re.match(r'^\d{1,2}-[A-Z]{3}-\d{2}$', clean_txt):
                                    continue
                                movimiento_actual["Descripción / Establecimiento"] += txt + " "

                if movimiento_actual:
                    for sp in skip_phrases:
                        movimiento_actual["Descripción / Establecimiento"] = re.sub(
                            re.escape(sp), "", movimiento_actual["Descripción / Establecimiento"], flags=re.IGNORECASE
                        )
                    todos_los_movimientos.append(movimiento_actual)

            df = pd.DataFrame(todos_los_movimientos, columns=[
                "Fecha",
                "Descripción / Establecimiento",
                "Monto del deposito",
                "Monto del retiro",
                "Saldo"
            ])

            ruta_salida = os.path.join(output_folder, excel_name)
            df.to_excel(ruta_salida, index=False)

            wb = load_workbook(ruta_salida)
            ws = wb.active

            ws.insert_rows(1, 6)

            ws["A1"] = f"Banco: Banorte"
            ws["A2"] = f"Empresa: {empresa_str}"
            ws["A3"] = f"No. Cliente: {no_cliente_str}"
            ws["A4"] = f"Periodo: {periodo_str}"
            ws["A5"] = f"RFC: {rfc_str}"

            thin_side = Side(border_style="thin")
            thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)

            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(1, max_col + 1):
                cell = ws.cell(row=7, column=col)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row in range(8, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[col_letter].width = max_length + 2

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            for sh in ("BDP1", "BDP2"):
                if sh in wb.sheetnames: del wb[sh]

            df=pd.read_excel(ruta_salida,engine="openpyxl")
            df_dep=df[df["Monto del retiro"].isna()].copy()    

            def volcar(ws_dest, modo):
                for i,col in enumerate(df_dep.columns,1):
                    h=ws_dest.cell(row=1,column=i,value=col)
                    h.font=Font(bold=True,color="FFFFFF"); h.fill=header_fill
                    h.alignment=Alignment(horizontal="center"); h.border=thin_border

                r=2
                for fecha,bloque in df_dep.groupby("Fecha"):
                # movimientos
                    for _,fila in bloque.iterrows():
                        for i,val in enumerate(fila,1):
                            c=ws_dest.cell(row=r,column=i,value=val); c.border=thin_border
                        r+=1
                    
                    bold=Font(bold=True)

                    if modo=="BDP1":
                        for _,fila in bloque.iterrows():
                            for i,val in enumerate(fila,1):
                                v="TOTAL" if i==1 else val
                                c=ws_dest.cell(row=r,column=i,value=v)
                                c.font=bold; c.border=thin_border
                            r+=1
                        r+=1  # blanco después

                    elif modo=="BDP2":
                        total=bloque["Monto del deposito"].sum()
                        ws_dest.cell(row=r,column=1,value="TOTAL").border=thin_border
                        ws_dest.cell(row=r,column=1,value="TOTAL").font=bold
                        ws_dest.cell(row=r,column=2,
                                    value=f"Ingresos totales del día {fecha}").border=thin_border
                        ws_dest.cell(row=r,column=2,
                                    value=f"Ingresos totales del día {fecha}").font=bold
                        ws_dest.cell(row=r,column=3,value=total).border=thin_border
                        ws_dest.cell(row=r,column=3,value=total).font=bold
                        r+=2  # blanco

                # auto‑ancho
                for col in ws_dest.columns:
                    ws_dest.column_dimensions[col[0].column_letter].width= \
                        max(len(str(c.value)) if c.value else 0 for c in col)+2
                
            volcar(wb.create_sheet("BDP1"), modo="BDP1")
            volcar(wb.create_sheet("BDP2"), modo="BDP2")
            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

def main():
    global pdf_paths, output_folder
    if len(sys.argv) < 2:
        return
    output_folder = sys.argv[1]
    pdf_paths = ""
    root = tk.Tk()
    root.title("Extracción Movimientos - Banorte")

    win_width = 600
    win_height = 250
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")
    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))

    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)

    global entry_archivo
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)

    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
    btn_procesar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
