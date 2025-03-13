import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# -------------------------------------------------
# Ajusta estos valores a tu PDF
LEFT_BOUND = 50.0       # Límite izquierdo para NO. REF.
RIGHT_BOUND = 90.0      # Límite derecho para NO. REF.
MIN_REF_FRACTION = 0.2  # Umbral mínimo de fracción en NO. REF.
# -------------------------------------------------

MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
                "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

def es_linea_movimiento(line_text):
    """
    Determina si la línea inicia un movimiento nuevo en formato 'dd mmm'.
    Ej: '14 ENE'
    """
    tokens = line_text.split()
    if len(tokens) < 2:
        return False
    if not re.match(r'^\d{1,2}$', tokens[0]):  # día
        return False
    if tokens[1] not in MESES_CORTOS:         # mes abreviado
        return False
    return True

def es_numero_monetario(texto):
    """
    Valida montos con/sin '$', con/sin 'USD' al final.
    Ej: 100,923.30 / $100,923.30 / $100,923.30 USD
    """
    patron = r'^\$?\s?[\d,]+\.\d{2}(?:\s?USD)?$'
    return bool(re.match(patron, texto.strip()))

def dist(a, b):
    """Distancia absoluta."""
    return abs(a - b)

def merge_tokens(token1, token2):
    """
    Une dos tokens (por ejemplo, '$' y '100.00') en uno solo,
    combinando sus coordenadas y textos.
    """
    new_text = token1['text'].strip() + token2['text'].strip()
    new_x0 = min(token1['x0'], token2['x0'])
    new_x1 = max(token1['x1'], token2['x1'])
    new_top = token1['top']
    merged = {
        'text': new_text,
        'x0': new_x0,
        'x1': new_x1,
        'top': new_top
    }
    return merged

def detect_headers(page0):
    """
    Recorre la primera página para detectar la posición (center_x) de:
      - NO. REF. (si buscas "DOCTO", etc.)
      - DESCRIPCION DE LA OPERACION
      - DEPOSITOS
      - RETIROS
      - SALDO (p.ej. en la 5ª aparición)
    """
    headers_to_find = {
        "NO. REF.": None,
        "DESCRIPCION DE LA OPERACION": None,
        "DEPOSITOS": None,
        "RETIROS": None,
        "SALDO": None
    }
    words_page0 = page0.extract_words()

    saldo_count = 0
    depositos_count = 0

    for w in words_page0:
        txt = w['text'].upper().strip()
        center_x = (w['x0'] + w['x1']) / 2

        # Ejemplo: si buscas "DOCTO" para NO. REF.:
        if txt == "DOCTO" and headers_to_find["NO. REF."] is None:
            headers_to_find["NO. REF."] = center_x

        elif "DESCRIPCION" in txt:
            if headers_to_find["DESCRIPCION DE LA OPERACION"] is None:
                headers_to_find["DESCRIPCION DE LA OPERACION"] = center_x

        elif "DEPOSITOS" in txt:
            depositos_count += 1
            if depositos_count == 2:
                headers_to_find["DEPOSITOS"] = center_x

        elif "RETIROS" in txt:
            if headers_to_find["RETIROS"] is None:
                headers_to_find["RETIROS"] = center_x

        elif "SALDO" in txt:
            saldo_count += 1
            # Ejemplo: a la 5ª vez
            if saldo_count == 5:
                headers_to_find["SALDO"] = center_x

    return headers_to_find

def cargar_archivo():
    global pdf_paths
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos

def procesar_pdf():
    global pdf_paths, output_folder
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return

    for pdf_path in pdf_paths:
        try:
            pdf_name = os.path.basename(pdf_path)
            pdf_stem, pdf_ext = os.path.splitext(pdf_name)
            excel_name = pdf_stem + ".xlsx"

            with pdfplumber.open(pdf_path) as pdf:
                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return

                # Detectar encabezados
                page0 = pdf.pages[0]
                header_positions = detect_headers(page0)

                mxn_movimientos = []
                usd_movimientos = []
                current_moneda = "MXN"

                skip_phrases = ["ESTADO DE CUENTA AL", "PÁGINA", "PAGINA", "CONTINUA EN LA SIGUIENTE PAGINA", "NO. REF. /", "DOCTO", "ESTADO DE CUENTA", "NUMERO DE CLIENTE:", "R.F.C."]
                skip_phrases = [s.upper() for s in skip_phrases]
                stop_phrases = ["SALDO TOTAL*"]
                stop_phrases = [s.upper() for s in stop_phrases]
                restart_phrase = ["DETALLE DE LA CUENTA: CUENTA DE CHEQUES EN DOLARES"]
                restart_phrase = [s.upper() for s in restart_phrase]

                start_reading = False
                stop_reading = False
                restart_phrase_found = False
                movimiento_actual = None

                # Datos de encabezado extra
                periodo_str = ""
                no_cuenta_str = ""
                empresa_str = ""
                no_cliente_str = ""
                rfc_str = ""

                for page_index, page in enumerate(pdf.pages):
                    if stop_reading:
                        break

                    words = page.extract_words()
                    if not words:
                        continue

                    # Agrupar tokens por 'top'
                    lineas_dict = {}
                    for w in words:
                        top_approx = int(w['top'])
                        lineas_dict.setdefault(top_approx, []).append(w)
                    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                    for top_val, words_in_line in lineas_ordenadas:
                        if stop_reading:
                            break

                        # Unir tokens tipo "$" + "100.00"
                        joined_tokens = []
                        i = 0
                        while i < len(words_in_line):
                            token = words_in_line[i]
                            txt = token['text'].strip()
                            if txt == "$" and (i + 1 < len(words_in_line)):
                                next_token = words_in_line[i+1]
                                combined_text = txt + next_token['text'].strip()
                                if es_numero_monetario(combined_text):
                                    merged = merge_tokens(token, next_token)
                                    joined_tokens.append(merged)
                                    i += 2
                                    continue
                                else:
                                    joined_tokens.append(token)
                                    i += 1
                            else:
                                joined_tokens.append(token)
                                i += 1

                        line_text = " ".join(t['text'].strip() for t in joined_tokens)
                        line_text_upper = line_text.upper()

                        # Omitir líneas con skip_phrases
                        if any(sp in line_text_upper for sp in skip_phrases):
                            continue
                        # Detener si aparece stop_phrases
                        if any(sp in line_text_upper for sp in stop_phrases):
                            stop_reading = True
                            break

                       
                        

                        # Detectar cambio de moneda
                        if "CUENTA DE CHEQUES EN DOLARES" in line_text_upper:
                            current_moneda = "USD"

                        # Capturar datos extra
                        if "RESUMEN" in line_text_upper and "DEL:" in line_text_upper:
                            tokens_line = line_text.split()
                            fechas = [t for t in tokens_line if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', t)]
                            if len(fechas) == 2:
                                periodo_str = f"{fechas[0]} al {fechas[1]}"
                            else:
                                periodo_str = line_text
                            continue
                        if "CONTRATO" in line_text_upper and not no_cuenta_str:
                            tokens_line = line_text.split()
                            no_cuenta_str = tokens_line[-1]
                            continue
                        if "NUMERO DE CLIENTE:" in line_text_upper and not no_cliente_str:
                            tokens_line = line_text.split()
                            no_cliente_str = tokens_line[-1]
                            continue
                        if "R.F.C." in line_text_upper and not rfc_str:
                            tokens_line = line_text.split()
                            rfc_str = tokens_line[-1]
                            continue

                        # Iniciar lectura de movimientos
                        if not start_reading:
                            if es_linea_movimiento(line_text_upper):
                                start_reading = True
                            else:
                                continue

                        # Detectar nueva línea de movimiento
                        if es_linea_movimiento(line_text_upper):
                            # Guardar movimiento anterior
                            if movimiento_actual:
                                if current_moneda == "USD":
                                    usd_movimientos.append(movimiento_actual)
                                else:
                                    mxn_movimientos.append(movimiento_actual)

                            # Creamos un nuevo movimiento con la fecha
                            tokens_line = line_text_upper.split()
                            nueva_fecha = f"{tokens_line[0]} {tokens_line[1]}"
                            movimiento_actual = {
                                "FECHA": nueva_fecha,
                                "NO. REF.": "",
                                "DESCRIPCION DE LA OPERACION": "",
                                "DEPOSITOS": None,
                                "RETIROS": None,
                                "SALDO": None
                            }

                            # Para no duplicar día/mes en descripción, filtramos
                            new_joined_tokens = []
                            day_found = False
                            month_found = False
                            for tk2 in joined_tokens:
                                t_up = tk2['text'].upper().strip()
                                if not day_found and re.match(r'^\d{1,2}$', t_up):
                                    day_found = True
                                    continue
                                elif day_found and not month_found and t_up in MESES_CORTOS:
                                    month_found = True
                                    continue
                                else:
                                    new_joined_tokens.append(tk2)

                            joined_tokens = new_joined_tokens

                        else:
                            if not movimiento_actual:
                                movimiento_actual = {
                                    "FECHA": None,
                                    "NO. REF.": "",
                                    "DESCRIPCION DE LA OPERACION": "",
                                    "DEPOSITOS": None,
                                    "RETIROS": None,
                                    "SALDO": None
                                }

                        # --- Asignar cada token ---
                        for tk in joined_tokens:
                            txt_joined = tk['text'].strip()
                            x0 = tk['x0']
                            x1 = tk['x1']

                            # 1) Montos monetarios
                            if es_numero_monetario(txt_joined):
                                center_w = (x0 + x1) / 2
                                dist_depositos = dist(center_w, header_positions["DEPOSITOS"]) if header_positions["DEPOSITOS"] else float('inf')
                                dist_retiros   = dist(center_w, header_positions["RETIROS"])   if header_positions["RETIROS"] else float('inf')
                                dist_saldo     = dist(center_w, header_positions["SALDO"])     if header_positions["SALDO"] else float('inf')
                                min_distance = min(dist_depositos, dist_retiros, dist_saldo)
                                if min_distance == dist_depositos:
                                    movimiento_actual["DEPOSITOS"] = txt_joined
                                elif min_distance == dist_retiros:
                                    movimiento_actual["RETIROS"] = txt_joined
                                elif min_distance == dist_saldo:
                                    movimiento_actual["SALDO"] = txt_joined

                            else:
                                # 2) Dividir token en 3 zonas: [x0, LEFT_BOUND], [LEFT_BOUND, RIGHT_BOUND], [RIGHT_BOUND, x1]
                                token_width = x1 - x0
                                if token_width <= 0:
                                    # Evitamos divisiones raras
                                    movimiento_actual["DESCRIPCION DE LA OPERACION"] += " " + txt_joined
                                    continue

                                # Fracción a la IZQUIERDA de LEFT_BOUND
                                fraction_left = 0.0
                                if x0 < LEFT_BOUND:
                                    overlap_left = min(LEFT_BOUND, x1) - x0
                                    if overlap_left < 0: 
                                        overlap_left = 0
                                    fraction_left = max(0, overlap_left / token_width)

                                # Fracción a la DERECHA de RIGHT_BOUND
                                fraction_right = 0.0
                                if x1 > RIGHT_BOUND:
                                    overlap_right = x1 - max(RIGHT_BOUND, x0)
                                    if overlap_right < 0:
                                        overlap_right = 0
                                    fraction_right = max(0, overlap_right / token_width)

                                # Fracción central (NO. REF.)
                                fraction_ref = 1 - fraction_left - fraction_right
                                if fraction_ref < 0:
                                    fraction_ref = 0

                                n = len(txt_joined)
                                cut_left = int(round(n * fraction_left))
                                cut_ref = cut_left + int(round(n * fraction_ref))

                                text_left = txt_joined[:cut_left]      # parte "izquierda"
                                text_ref  = txt_joined[cut_left:cut_ref] # parte "NO. REF."
                                text_right= txt_joined[cut_ref:]       # parte "derecha"

                                # --- Aplicar umbral mínimo para NO. REF. ---
                                # Si fraction_ref < MIN_REF_FRACTION, se descarta la parte central
                                if fraction_ref < MIN_REF_FRACTION:
                                    # Se manda el token entero a la zona con mayor fracción
                                    # (Por simplicidad, lo mandamos a DESCRIPCION, 
                                    #  pero podrías comparar fraction_left vs fraction_right)
                                    movimiento_actual["DESCRIPCION DE LA OPERACION"] += " " + txt_joined
                                    continue

                                # Asignar la parte "left" a DESCRIPCION
                                if text_left:
                                    movimiento_actual["DESCRIPCION DE LA OPERACION"] += " " + text_left

                                # Asignar la parte "ref" a NO. REF.
                                if text_ref:
                                    if movimiento_actual["NO. REF."]:
                                        movimiento_actual["NO. REF."] += " " + text_ref
                                    else:
                                        movimiento_actual["NO. REF."] = text_ref

                                # Asignar la parte "right" a DESCRIPCION
                                if text_right:
                                    movimiento_actual["DESCRIPCION DE LA OPERACION"] += " " + text_right

                # Guardar último movimiento
                if movimiento_actual:
                    if current_moneda == "USD":
                        usd_movimientos.append(movimiento_actual)
                    else:
                        mxn_movimientos.append(movimiento_actual)

            # Crear DataFrames
            df_mxn = pd.DataFrame(mxn_movimientos, columns=[
                "FECHA", "NO. REF.", "DESCRIPCION DE LA OPERACION", "DEPOSITOS", "RETIROS", "SALDO"
            ])
            df_usd = pd.DataFrame(usd_movimientos, columns=[
                "FECHA", "NO. REF.", "DESCRIPCION DE LA OPERACION", "DEPOSITOS", "RETIROS", "SALDO"
            ])

            # Guardar a Excel
            ruta_salida = os.path.join(output_folder, excel_name)
            with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
                df_mxn.to_excel(writer, sheet_name="Movs_MXN", index=False)
                df_usd.to_excel(writer, sheet_name="Movs_USD", index=False)

                wb = writer.book
                ws_mxn = wb["Movs_MXN"]
                ws_usd = wb["Movs_USD"]

                # Encabezado extra en la hoja MXN
                ws_mxn.insert_rows(1, 6)
                ws_mxn["A1"] = f"Banco: BanBajío"
                ws_mxn["A2"] = f"Empresa: {empresa_str}"
                ws_mxn["A3"] = f"No. Cuenta: {no_cuenta_str}"
                ws_mxn["A4"] = f"No. Cliente: {no_cliente_str}"
                ws_mxn["A5"] = f"Periodo: {periodo_str}"
                ws_mxn["A6"] = f"RFC: {rfc_str}"

                # Encabezado extra en la hoja USD
                ws_usd.insert_rows(1, 6)
                ws_usd["A1"] = f"Banco: BanBajío"
                ws_usd["A2"] = f"Empresa: {empresa_str}"
                ws_usd["A3"] = f"No. Cuenta: {no_cuenta_str}"
                ws_usd["A4"] = f"No. Cliente: {no_cliente_str}"
                ws_usd["A5"] = f"Periodo: {periodo_str}"
                ws_usd["A6"] = f"RFC: {rfc_str}"

                # Formato de celdas
                thin_side = Side(border_style="thin")
                thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True)

                # Formatear Movs_MXN
                max_row_mxn = ws_mxn.max_row
                max_col_mxn = ws_mxn.max_column
                for col in range(1, max_col_mxn + 1):
                    cell = ws_mxn.cell(row=7, column=col)
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                for row in range(8, max_row_mxn + 1):
                    for col in range(1, max_col_mxn + 1):
                        cell = ws_mxn.cell(row=row, column=col)
                        cell.border = thin_border
                for col in ws_mxn.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    ws_mxn.column_dimensions[col_letter].width = max_length + 2
                for row in ws_mxn.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                # Formatear Movs_USD
                max_row_usd = ws_usd.max_row
                max_col_usd = ws_usd.max_column
                for col in range(1, max_col_usd + 1):
                    cell = ws_usd.cell(row=7, column=col)
                    cell.fill = header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                for row in range(8, max_row_usd + 1):
                    for col in range(1, max_col_usd + 1):
                        cell = ws_usd.cell(row=row, column=col)
                        cell.border = thin_border
                for col in ws_usd.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            length = len(str(cell.value))
                            if length > max_length:
                                max_length = length
                    ws_usd.column_dimensions[col_letter].width = max_length + 2
                for row in ws_usd.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                wb.save(ruta_salida)

            messagebox.showinfo("Éxito", f"Archivo Excel generado con 2 hojas: {ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

def main():
    global pdf_paths, output_folder
    if len(sys.argv) < 2:
        return
    output_folder = sys.argv[1]
    pdf_paths = ""

    root = tk.Tk()
    root.title("Extracción Movimientos - banamex")
    win_width = 600
    win_height = 250

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")

    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))

    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)

    global entry_archivo
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)

    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
    btn_procesar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
