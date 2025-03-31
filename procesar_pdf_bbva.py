import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side



def es_fecha_valida(texto):
    """
    Verifica si algo como '2/ENE' o '15/FEB' coincide con el patrón de fecha.
    """
    patron = r'^\d{1,2}/[A-Z]{3}$'
    return bool(re.match(patron, texto.strip()))

def es_linea_movimiento(linea):
    """
    Determina si la línea inicia un 'movimiento' nuevo.
    Regresará True si los primeros 2 'tokens' son fechas tipo dd/ENE.
    """
    tokens = linea.split()
    if len(tokens) < 2:
        return False
    return es_fecha_valida(tokens[0]) and es_fecha_valida(tokens[1])

def es_numero_monetario(texto):
    """
    Determina si un texto es un número tipo '100,923.30'.
    Ajusta si tu PDF usa otro formato (p.ej. 100.923,30).
    """
    return bool(re.match(r'^[\d,]+\.\d{2}$', texto.strip()))

def parse_monetario(txt):
    """
    Convierte un texto como '100,923.30' a un float.
    """
    txt = txt.strip()
    sign = 1

    # Verifica si está entre paréntesis (se asume negativo)
    if txt.startswith("(") and txt.endswith(")"):
        sign = -1
        txt = txt[1:-1].strip()
    # Verifica si empieza con signo "-"
    elif txt.startswith("-"):
        sign = -1
        txt = txt[1:].strip()

    # Elimina comas antes de convertir a float
    txt = txt.replace(",", "")
    return sign * float(txt)

def dist(a, b):
    """Distancia absoluta entre dos valores."""
    return abs(a - b)

def cargar_archivo():
    global pdf_paths
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos


def procesar_pdf():
    global pdf_paths, output_folder
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return

    for pdf_path in pdf_paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:



                pdf_name = os.path.basename(pdf_path)
                pdf_stem, pdf_ext = os.path.splitext(pdf_name)
                excel_name = pdf_stem + ".xlsx"

                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return

                periodo_str = ""
                no_cuenta_str = ""
                empresa_str = ""
                no_cliente_str = ""
                rfc_str = ""

                # 1) Detectar las posiciones X de los encabezados en la primera página
                page0_words = pdf.pages[0].extract_words()
                col_positions = {}  # dict { "CARGOS": x_center, "ABONOS": x_center, ... }

                regionEmpresa = (30, 82.69857999999999, 242.925, 92.69857999999999)

                croppedEmpresa = pdf.pages[0].within_bbox(regionEmpresa)
                empresa_str = croppedEmpresa.extract_text() or ""
                print(empresa_str)

                # Ajusta estos nombres según tu PDF:
                encabezados_buscar = ["CARGOS", "ABONOS", "OPERACIÓN", "LIQUIDACIÓN"]

                for w in page0_words:
                    txt_upper = w['text'].strip().upper()
                    center_x = (w['x0'] + w['x1']) / 2
                    # print(f"Texto: {w['text']}, x0: {w['x0']}, x1: {w['x1']}, top: {w['top']}, bottom: {w['bottom']}")
                    if txt_upper in encabezados_buscar:
                        col_positions[txt_upper] = center_x

                # Si no detectaste todos, podrías asignar manualmente:
                # col_positions["CARGOS"] = 350
                # col_positions["ABONOS"] = 420
                # col_positions["SALDO OPERACIÓN"] = 490
                # col_positions["SALDO LIQUIDACIÓN"] = 560

                # Ordenamos las columnas por su x_center
                columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])
                # columnas_ordenadas = [("CARGOS", 350), ("ABONOS", 420), ...]

                # 2) Definir skip_phrases, stop_phrases, etc.
                skip_phrases = [
                    "Ciudad de México", "Av. Paseo de la Reforma", "R.F.C.",
                    "La GAT Real", "Información Financiera", "SUCURSAL", "DIRECCION",
                    "PLAZA", "TELEFONO", "N/A", "Saldo Promedio", "Rendimiento",
                    "Comisiones de la cuenta", "Cargos Objetados", "Saldo de Liquidación Inicial",
                    "Saldo Final", "Estimado Cliente", "Estado de Cuenta", "MAESTRA",
                    "PAGINA", "No. Cuenta", "No. Cliente", "FECHA", "SALDO", "OPER",
                    "LIQ", "COD.", "DESCRIPCION", "REFERENCIA", "CARGOS", "ABONOS",
                    "OPERACION", "LIQUIDACION", "También le informamos", "el cual puede",
                    "Con BBVA", "BBVA MEXICO, S.A."
                ]
                stop_phrases = ["Total de Movimientos"]

                start_reading = False
                stop_reading = False

                todos_los_movimientos = []
                movimiento_actual = None

                # 3) Recorremos todas las páginas
                for page_index, page in enumerate(pdf.pages):
                    if stop_reading:
                        break

                    # Extraemos las words con sus coordenadas
                    words = page.extract_words()

                    # Si es la primera página, imprimimos en consola las posiciones
                    #if page_index == 0:
                    #   words_debug = page.extract_words()
                    #  for w in words_debug:
                    #     print(f"[Página {page_index+1}] Texto: '{w['text']}' -> "
                        #        f"x0: {w['x0']}, x1: {w['x1']}, top: {w['top']}, bottom: {w['bottom']}")
                    #else:
                    #   words_debug = page.extract_words()


                    # Agrupamos por 'top' aproximado para formar líneas
                    lineas_dict = {}
                    for w in words:
                        top_approx = int(w['top'])  # redondeamos
                        if top_approx not in lineas_dict:
                            lineas_dict[top_approx] = []
                        lineas_dict[top_approx].append(w)

                    # Ordenamos las líneas de arriba hacia abajo
                    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                    for top_val, words_in_line in lineas_ordenadas:
                        if stop_reading:
                            break

                        # Convertimos la línea a string (para skip_phrases, etc.)
                        line_text = " ".join(w['text'] for w in words_in_line)
                        
                        # revisar el rango para extraer la info del inicio
                        # periodo => 56.80
                            # 4) Detectar "Periodo" si quieres
                        if "Periodo" in line_text and "DEL" in line_text:
                            # Ej: "Periodo DEL 01/01/2024 AL 31/01/2024"
                            tokens = line_text.split()
                            # busca tokens que coincidan con dd/mm/yyyy
                            fechas = [t for t in tokens if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', t)]
                            if len(fechas) == 2:
                                periodo_str = f"{fechas[0]} al {fechas[1]}"
                            else:
                                periodo_str = line_text
                            continue
        

                        if "No. de Cuenta" in line_text and not no_cuenta_str:
                            # line_text = "No. de Cuenta 0156337112"
                            tokens = line_text.split()  # ["No.", "de", "Cuenta", "0156337112"]
                            no_cuenta_str = tokens[-1]  # "0156337112"

                        # Empresa => ~94.03
                        # if 94.00 <= top_val <= 94.06:
                        #     # "TRAFFICLIGHT DE MEXICO SA DE CV"
                        #     empresa_str = line_text
                        #     continue

                        if "No. de Cliente" in line_text and not no_cliente_str:
                            tokens = line_text.split()
                            no_cliente_str = tokens[-1]
                            continue    

                        if "R.F.C" in line_text and not rfc_str:
                            tokens = line_text.split()
                            rfc_str = tokens[-1]
                            continue        

                        # Checar stop_phrases
                        if any(sp in line_text for sp in stop_phrases):
                            stop_reading = True
                            break

                        # Hasta que no encontremos la primera fecha, ignoramos
                        if not start_reading:
                            tokens = line_text.split()
                            found_date = any(es_fecha_valida(t) for t in tokens)
                            if found_date:
                                start_reading = True
                            else:
                                continue

                        # skip_phrases
                        if any(sp in line_text for sp in skip_phrases):
                            continue

                        # 4) Ver si la línea inicia un nuevo movimiento
                        if es_linea_movimiento(line_text):
                            # Cerramos el movimiento anterior si existe
                            if movimiento_actual:
                                todos_los_movimientos.append(movimiento_actual)

                            # Creamos un nuevo dict para el movimiento
                            tokens_line = line_text.split()
                            movimiento_actual = {
                                "Fecha operación": tokens_line[0],
                                "Fecha liquidación": tokens_line[1],
                                "Con. Descripción": "",
                                "Referencia": None,
                                "Cargos": None,
                                "Abonos": None,
                                "Saldo operación": None,
                                "Saldo liquidación": None
                            }

                            # Quitamos esos 2 tokens de la línea para que no estorben
                            # y luego asignaremos montos a columnas
                            # tokens_line = tokens_line[2:]
                            # Pero usaremos words_in_line para la asignación con x0

                        else:
                            # Es continuación
                            if not movimiento_actual:
                                movimiento_actual = {
                                    "Fecha operación": None,
                                    "Fecha liquidación": None,
                                    "Con. Descripción": "",
                                    "Referencia": None,
                                    "Cargos": None,
                                    "Abonos": None,
                                    "Saldo operación": None,
                                    "Saldo liquidación": None
                                }

                        # 5) Asignar montos por coordenadas
                        for w in words_in_line:
                            txt = w['text'].strip()
                            center_w = (w['x0'] + w['x1']) / 2

                            # Si es un número monetario, lo ubicamos en la columna más cercana
                            if es_numero_monetario(txt):
                                val = parse_monetario(txt)
                                if columnas_ordenadas:
                                    col_name, col_center = min(
                                        columnas_ordenadas,
                                        key=lambda x: dist(x[1], center_w)
                                    )
                                    # col_name es algo como "CARGOS", "ABONOS", etc.
                                    if col_name == "CARGOS":
                                        movimiento_actual["Cargos"] = val
                                    elif col_name == "ABONOS":
                                        movimiento_actual["Abonos"] = val
                                    elif col_name == "OPERACIÓN":
                                        movimiento_actual["Saldo operación"] = val
                                    elif col_name == "LIQUIDACIÓN":
                                        movimiento_actual["Saldo liquidación"] = val
                                else:
                                    # Si no detectamos encabezados, por defecto a "Cargos"
                                    movimiento_actual["Cargos"] = val
                            else:
                                # Si no es monetario y no es la fecha de la línea,
                                # considerarlo parte de la descripción
                                # Evitamos duplicar las 2 fechas iniciales
                                if es_fecha_valida(txt):
                                    continue
                                movimiento_actual["Con. Descripción"] += " " + txt

                # Al terminar TODAS las páginas
                if movimiento_actual:
                    todos_los_movimientos.append(movimiento_actual)

            # print("DEBUG - Datos capturados:")
            # print(f"Periodo: {periodo_str}")
            # print(f"No. de Cuenta: {no_cuenta_str}")
            # print(f"Empresa: {empresa_str}")
            # print(f"No. de Cliente: {no_cliente_str}")

            # Convertir a DataFrame
            df = pd.DataFrame(todos_los_movimientos, columns=[
                "Fecha operación",
                "Fecha liquidación",
                "Con. Descripción",
                "Referencia",
                "Cargos",
                "Abonos",
                "Saldo operación",
                "Saldo liquidación"
            ])

            # ruta_salida = os.path.join(output_folder,"movimientos_bbva_dd_mmm.xlsx")
            ruta_salida = os.path.join(output_folder, excel_name)
            df.to_excel(ruta_salida, index=False)

            # Ajustar ancho de columnas con openpyxl
            wb = load_workbook(ruta_salida)
            ws = wb.active

            # insertar 5 filas arriba
            ws.insert_rows(1, 6)

            # Añadir título
            ws["A1"] = f"Banco: BBVA México"
            ws["A2"] = f"Empresa: {empresa_str}"
            ws["A3"] = f"No. Cuenta: {no_cuenta_str}"
            ws["A4"] = f"No. Cliente: {no_cliente_str}"  
            ws["A5"] = f"Periodo: {periodo_str}"
            ws["A6"] = f"RFC: {rfc_str}"


            #Bordes
            thin_side = Side(border_style="thin")
            thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

            #color de fondo para fila encabezados
            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)

            # Max de filas y columnas
            max_row = ws.max_row
            max_col = ws.max_column

            # 1) Estilamos la fila de encabezados de la tabla (que está en la fila 7)
            for col in range(1, max_col+1):
                cell = ws.cell(row=7, column=col)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # 2) Estilamos las filas de datos (desde la 8 hasta max_row)
            for row in range(8, max_row+1):
                for col in range(1, max_col+1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    # Podrías poner alignment a la derecha para montos, etc.
                    # if col in [5, 6, 7, 8]:  # Cargos, Abonos, Saldo Op, Saldo Liq
                    #     cell.alignment = Alignment(horizontal="right")
    


            # # Ajustar estilos
            # bold_font = Font(bold=True)
            # for fila_encabezado in range(1, 7):
            #     cell = ws.cell(row=fila_encabezado, column=1)
            #     cell.font = bold_font

            #     cell.alignment = Alignment(horizontal="left")

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[col_letter].width = max_length + 2

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

def main():
    global pdf_paths, output_folder
    if len(sys.argv) < 2:
        print("Uso: python procesar_pdf_bbva.py <carpeta_salida>")
        return
    output_folder = sys.argv[1]

    pdf_paths = ""
    # Interfaz gráfica con tkinter
    root = tk.Tk()

    root.title("Extracción Movimientos - bbva")
    win_width = 600
    win_height = 250

    # Obtenemos dimensiones de la pantalla
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

        # Calculamos coordenadas x e y
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    # Ajustamos la geometría: ancho x alto + x + y
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")

    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))

    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)

    global entry_archivo
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)

    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
    btn_procesar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
