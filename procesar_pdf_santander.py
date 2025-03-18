import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# -------------------------------------------------
# Ajusta estos valores a tu PDF
LEFT_BOUND = 60.229       # Límite izquierdo para NO. REF.
RIGHT_BOUND = 102.0       # Límite derecho para NO. REF.
MIN_REF_FRACTION = 0.2    # Umbral mínimo de fracción en NO. REF.
# -------------------------------------------------

MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
                "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

def es_linea_movimiento(linea):
    """
    Determina si la línea (o token) inicia un 'movimiento' nuevo
    en formato 'dd-mmm-aaaa', donde:
      - dd: día (1-2 dígitos)
      - mmm: mes (3 letras mayúsculas, ej. DIC, ENE)
      - aaaa: año (4 dígitos)
    """
    match = re.match(r'^(\d{1,2})-([A-Z]{3})-(\d{4})$', linea.strip())
    if not match:
        return False
    dia, mes, anio = match.groups()
    if mes not in MESES_CORTOS:
        return False
    return True

def es_numero_monetario(texto):
    """
    Determina si un texto es un número tipo '100,923.30'.
    Ajusta si tu PDF usa otro formato (p.ej. 100.923,30).
    """
    return bool(re.match(r'^[\d,]+\.\d{2}$', texto.strip()))

def dist(a, b):
    """Distancia absoluta entre dos valores."""
    return abs(a - b)

def cargar_archivo():
    global pdf_paths
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos

def procesar_pdf():
    global pdf_paths, output_folder
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return

    for pdf_path in pdf_paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                pdf_name = os.path.basename(pdf_path)
                pdf_stem, pdf_ext = os.path.splitext(pdf_name)
                excel_name = pdf_stem + ".xlsx"

                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return

                # =======================
                # 1) DETECTAR ENCABEZADOS EN LA 1RA PÁGINA
                # =======================
                page0 = pdf.pages[0]

                regionEmpresa = (43.5, 71.729, 369, 76.63)
                regionNoCliente = (479.98, 77.378, 528, 82.279)
 

                croppedEmpresa = page0.within_bbox(regionEmpresa)
                croppedNoCliente = page0.within_bbox(regionNoCliente)
                empresa_str = croppedEmpresa.extract_text() or ""
                no_cliente_str = croppedNoCliente.extract_text() or ""

                words_page0 = page0.extract_words()

                encabezados_buscar = ["RETIROS", "DEPOSITOS", "SALDO"]
                col_positions = {}

                # Agrupamos las palabras de la primera página por 'top' para formar líneas
                lineas_dict_page0 = {}
                for w in words_page0:
                    print(f"Texto: {w['text']}, x0: {w['x0']}, x1: {w['x1']}, top: {w['top']}, bottom: {w['bottom']}")
                    top_approx = int(w['top'])
                    if top_approx not in lineas_dict_page0:
                        lineas_dict_page0[top_approx] = []
                    lineas_dict_page0[top_approx].append(w)

                lineas_ordenadas_page0 = sorted(lineas_dict_page0.items(), key=lambda x: x[0])

                # Buscamos la línea que contenga los 3 encabezados
                for top_val, words_in_line in lineas_ordenadas_page0:
                    line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
                    if all(h in line_text_upper for h in encabezados_buscar):
                        for w in words_in_line:
                            w_text_upper = w['text'].strip().upper()
                            if w_text_upper in encabezados_buscar:
                                center_x = (w['x0'] + w['x1']) / 2
                                col_positions[w_text_upper] = center_x
                        break

                columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])
                print("columnas_ordenadas =", columnas_ordenadas)

                # =======================
                # 2) VARIABLES PARA ENCABEZADOS DEL EXCEL
                # =======================
                periodo_str = ""
                no_cuenta_str = ""
                # empresa_str = ""
                # no_cliente_str = ""
                rfc_str = ""

                # =======================
                # 3) FRASES A OMITIR (skip) Y A DETENER (stop)
                # =======================
                skip_phrases = [
                    "ESTADO DE CUENTA AL",
                    "PÁGINA"
                ]
                skip_phrases = [s.upper() for s in skip_phrases]

                stop_phrases = [
                    "INFORMACIONFISCAL",
                ]

                # ** Header repetido ** (el que quieres omitir si aparece en cada página)
                # Ajusta o añade más strings según el texto exacto que quieras saltar.
                header_phrases = [
                    "F E C H A FOLIO DESCRIPCION DEPOSITOS RETIROS SALDO",
                    "FECHA FOLIO DESCRIPCION DEPOSITOS RETIROS SALDO",
                    # etc...
                ]

                # ** Footer phrases ** (para cortar la lectura en esa página)
                footer_phrases = [
                    "BANCO SANTANDER (MEXICO)",
                    "INSTITUCION DE BANCA MULTIPLE,"
                ]

                start_reading = False
                stop_reading = False

                todos_los_movimientos = []
                movimiento_actual = None

                # =======================
                # 4) RECORRER TODAS LAS PÁGINAS
                # =======================
                for page_index, page in enumerate(pdf.pages):
                    if stop_reading:
                        break

                    words = page.extract_words()
                    lineas_dict = {}
                    for w in words:
                        top_approx = int(w['top'])
                        if top_approx not in lineas_dict:
                            lineas_dict[top_approx] = []
                        lineas_dict[top_approx].append(w)

                    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                    for top_val, words_in_line in lineas_ordenadas:
                        if stop_reading:
                            break

                        line_text = " ".join(w['text'] for w in words_in_line)
                        line_text_upper = line_text.upper().strip()

                        if "R.F.C." in line_text_upper and not rfc_str:
                            tokens_line = line_text.split()

                            if "R.F.C." in tokens_line:
                                rfc_idx = tokens_line.index("R.F.C.")
                                rfc_str = " ".join(tokens_line[rfc_idx + 1:])
                            # rfc_str = tokens_line[-1]
                            continue
                        if "PERIODO :" in line_text_upper and not periodo_str:
                            tokens_line = line_text.split()

                            if "PERIODO" in tokens_line:
                                periodo_idx = tokens_line.index("PERIODO")
                                periodo_str = " ".join(tokens_line[periodo_idx + 1:])
                            # rfc_str = tokens_line[-1]
                            continue
                        # 1) Footer check => si encontramos algo del footer, cortamos la lectura de la página
                        if any(fp in line_text_upper for fp in footer_phrases):
                            print("Footer detectado en la página. Omitimos el resto de la página.")
                            break  # Rompe el for => salta a la siguiente página

                        # 2) Si es la línea final de totales => la omitimos
                        if line_text_upper.startswith("TOTAL"):
                            tokens_line = line_text.split()
                            montos_en_linea = [t for t in tokens_line if es_numero_monetario(t)]
                            if len(montos_en_linea) > 0:
                                continue  # omitir esta línea

                        # 3) Revisar stop_phrases (para dejar de leer PDF por completo)
                        if any(sp in line_text_upper for sp in stop_phrases):
                            stop_reading = True
                            break

                        # 4) Activar lectura al encontrar la primera fecha
                        if not start_reading:
                            tokens_line = line_text.split()
                            found_date_token = any(es_linea_movimiento(t.upper()) for t in tokens_line)
                            if found_date_token:
                                start_reading = True
                            else:
                                continue

                        # 5) Omitir líneas con skip_phrases
                        if any(sp in line_text_upper for sp in skip_phrases):
                            continue

                        # 6) Omitir header repetido si coincide con header_phrases
                        if any(hp in line_text_upper for hp in header_phrases):
                            print("Header repetido detectado. Omitimos esta línea.")
                            continue

                        # 7) ¿Es un nuevo movimiento? => si la línea empieza con dd-mmm-aaaa
                        tokens_line = line_text_upper.split()
                        if len(tokens_line) > 0 and es_linea_movimiento(tokens_line[0]):
                            # Guardar el anterior, si existía
                            if movimiento_actual:
                                todos_los_movimientos.append(movimiento_actual)

                            movimiento_actual = {
                                "Fecha": tokens_line[0],
                                "Folio": None,
                                "Descripción": "",
                                "Depositos": None,
                                "Retiros": None,
                                "Saldo": None,
                                "Fecha_tokens": [tokens_line[0]]
                            }
                        else:
                            if not movimiento_actual:
                                movimiento_actual = {
                                    "Fecha": None,
                                    "Folio": None,
                                    "Descripción": "",
                                    "Depositos": None,
                                    "Retiros": None,
                                    "Saldo": None
                                }

                        # 8) Asignar folio, montos y descripción
                        for w in words_in_line:
                            txt = w['text'].strip()
                            center_w = (w['x0'] + w['x1']) / 2

                            # a) Folio
                            if (movimiento_actual is not None 
                                and movimiento_actual.get("Folio") is None
                                and LEFT_BOUND <= center_w <= RIGHT_BOUND):
                                if txt not in movimiento_actual.get("Fecha_tokens", []):
                                    if re.match(r'^\d+$', txt):
                                        movimiento_actual["Folio"] = txt
                                        continue

                            # b) Número monetario
                            if es_numero_monetario(txt):
                                if columnas_ordenadas:
                                    col_name, col_center = min(
                                        columnas_ordenadas,
                                        key=lambda x: dist(x[1], center_w)
                                    )
                                    if col_name == "DEPOSITOS":
                                        movimiento_actual["Depositos"] = txt
                                    elif col_name == "RETIROS":
                                        movimiento_actual["Retiros"] = txt
                                    elif col_name == "SALDO":
                                        movimiento_actual["Saldo"] = txt
                                else:
                                    movimiento_actual["Depositos"] = txt
                            else:
                                # c) Descripción
                                if txt not in movimiento_actual.get("Fecha_tokens", []):
                                    movimiento_actual["Descripción"] += " " + txt

                    # fin de la iteración de líneas en la página

                # Al terminar todas las páginas
                if movimiento_actual:
                    todos_los_movimientos.append(movimiento_actual)

            # =======================
            # 5) GUARDAR EN EXCEL
            # =======================
            df = pd.DataFrame(todos_los_movimientos, columns=[
                "Fecha",
                "Folio",
                "Descripción",
                "Depositos",
                "Retiros",
                "Saldo"
            ])

            ruta_salida = os.path.join(output_folder, excel_name)
            df.to_excel(ruta_salida, index=False)

            # Ajustes de estilo con openpyxl
            wb = load_workbook(ruta_salida)
            ws = wb.active

            ws.insert_rows(1, 6)
            ws["A1"] = f"Banco: Santander"
            ws["A2"] = f"Empresa: {empresa_str}"
            ws["A3"] = f"No. Cuenta: {no_cuenta_str}"
            ws["A4"] = f"No. Cliente: {no_cliente_str}"
            ws["A5"] = f"Periodo: {periodo_str}"
            ws["A6"] = f"RFC: {rfc_str}"

            thin_side = Side(border_style="thin")
            thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)

            max_row = ws.max_row
            max_col = ws.max_column

            # Estilo para la fila de encabezados (fila 7)
            for col in range(1, max_col + 1):
                cell = ws.cell(row=7, column=col)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Estilo para filas de datos
            for row in range(8, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[col_letter].width = max_length + 2

            # Ajuste wrap_text
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

def main():
    global pdf_paths, output_folder
    if len(sys.argv) < 2:
        print("Uso: python procesar_pdf_banamex.py <carpeta_salida>")
        return
    output_folder = sys.argv[1]
    pdf_paths = ""

    root = tk.Tk()
    root.title("Extracción Movimientos - banamex")
    win_width = 600
    win_height = 250

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")
    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))

    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)

    global entry_archivo
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)

    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
    btn_procesar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()
