import sys, os, re, tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ───────────────────── CONFIGURACIÓN GENERAL ──────────────────────
MESES = {"ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO",
         "SEP","OCT","NOV","DIC"}
DEBUG = True                              # imprime tokens en consola
DESCRIP_X0_MIN, DESCRIP_X0_MAX = 113.9, 350.0

# altura (bottom – top) característica del token TOTAL
TOTAL_H_MIN = 9.0         # ajusta si ves variaciones
TOTAL_H_MAX = 11.0
# ───────────────────────────────────────────────────────────────────


# ---------------------------- utilidades ---------------------------
def es_fecha(tok:str)->bool:
    return bool(re.match(r'^\d{2}-[A-Z]{3}-\d{4}$', tok)) and tok[3:6] in MESES

def es_monto(t:str)->bool:
    return bool(re.match(r'^[\d,]+\.\d{2}$', t.strip()))

def monto_float(t:str)->float:
    t=t.strip().replace(',', '')
    sign=-1 if t.startswith('(') or t.startswith('-') else 1
    t=t.strip('()-')
    return sign*float(t)

def agrupar_top(page):
    d={}
    for w in page.extract_words():
        d.setdefault(int(w['top']), []).append(w)
    return d

def detectar_columnas(pdf):
    buscados={"DEPOSITO","RETIRO","SALDO"}
    for p in pdf.pages:
        for _, ws in sorted(agrupar_top(p).items()):
            linea=" ".join(w['text'].upper() for w in ws)
            if all(x in linea for x in buscados):
                return {w['text'].upper(): (w['x0']+w['x1'])/2
                        for w in ws if w['text'].upper() in buscados}
    return {}
# -------------------------------------------------------------------


# -------------------- lógica principal de proceso ------------------
def procesar_archivos(pdf_paths, output_folder):
    for path in pdf_paths:
        with pdfplumber.open(path) as pdf:
            # depuración opcional de la página 11 (índice 10)
            if DEBUG and len(pdf.pages) > 10:
                print(f"\nDEBUG {os.path.basename(path)}  página 11 (índice 10)")
                for w in pdf.pages[10].extract_words():
                    alto=w['bottom']-w['top']
                    print(f"Texto:{w['text']!r}  x0={w['x0']}  top={w['top']}  h={alto:.2f}")

            cols = detectar_columnas(pdf)

            # ── encabezado (pág. 1) ──
            txt0 = pdf.pages[0].extract_text() or ""
            empresa=no_cli=periodo=rfc=""
            for ln in txt0.splitlines():
                u=ln.upper()
                if "CODIGO DE CLIENTE NO." in u and not no_cli:
                    no_cli=u.split("CODIGO DE CLIENTE NO.")[1].strip()
                if "R.F.C." in u and not rfc:
                    rfc=u.split("R.F.C.")[1].strip()
                if "PERIODO DEL" in u and not periodo:
                    m=re.search(r"PERIODO DEL (.+)", ln, re.I)
                    if m: periodo=m.group(1).strip()
                if "DESARROLLADORA" in u and not empresa:
                    empresa=ln.strip()

            movimientos=[]; start=False; mov=None; stop_reading=False

            for pg in pdf.pages[1:]:
                if stop_reading:
                    break
                for _, ws in sorted(agrupar_top(pg).items()):
                    if stop_reading:
                        break

                    linea=" ".join(w['text'].upper() for w in ws)
                    if "FECHA FOLIO DESCRIPCION" in linea: continue
                    if any(x in linea for x in ("ESTADO DE CUENTA","PÁGINA")): continue

                    toks=linea.split()
                    if toks and not start and es_fecha(toks[0][:11]):
                        start=True
                    if not start: continue

                    # ¿nueva línea con fecha?
                    if toks and es_fecha(toks[0][:11]):
                        if mov: movimientos.append(mov)
                        mov={"Fecha": toks[0][:11], "Folio": None,
                             "Descripción":"", "Depositos":None,
                             "Retiros":None, "Saldo":None,
                             "Fecha_tokens":[toks[0][:11]]}

                    if not mov: continue

                    for w in ws:
                        txt=w['text'].strip()

                        # ---- DETENER EN TOTAL con altura característica ----
                        alto = w['bottom'] - w['top']
                        if txt.upper() == "TOTAL" and TOTAL_H_MIN <= alto <= TOTAL_H_MAX:
                            stop_reading = True
                            break
                        # ----------------------------------------------------

                        cx=(w['x0']+w['x1'])/2

                        # token fusionado con fecha
                        if re.match(r'^\d{2}-[A-Z]{3}-\d{4}', txt) and len(txt)>11:
                            fecha=txt[:11]
                            if es_fecha(fecha):
                                mov["Fecha"]=fecha
                                resto=txt[11:].strip()
                                m=re.match(r'^(\d+)(.*)', resto)
                                if m:
                                    mov["Folio"]=m.group(1)
                                    resto=m.group(2).strip()
                                if resto: mov["Descripción"]+=" "+resto
                            continue

                        # folio suelto
                        if mov["Folio"] is None and txt.isdigit():
                            mov["Folio"]=txt; continue

                        # descripción
                        if DESCRIP_X0_MIN <= w['x0'] <= DESCRIP_X0_MAX:
                            mov["Descripción"]+=" "+txt; continue

                        # montos
                        if es_monto(txt):
                            val=monto_float(txt)
                            if cols:
                                col=min(cols, key=lambda c:abs(cols[c]-cx))
                                if col=="DEPOSITO": mov["Depositos"]=val
                                elif col=="RETIRO": mov["Retiros"]=val
                                else: mov["Saldo"]=val
                            else:
                                mov["Depositos"]=val
                        else:
                            mov["Descripción"]+=" "+txt
                # fin bucle líneas
            # fin bucle páginas
            if mov and not stop_reading:
                movimientos.append(mov)

        # ── guardar Excel ──
        df=pd.DataFrame(movimientos,
            columns=["Fecha","Folio","Descripción","Depositos","Retiros","Saldo"])
        out_xls=os.path.join(output_folder,
            os.path.splitext(os.path.basename(path))[0]+".xlsx")
        df.to_excel(out_xls,index=False)

        # formato rápido
        wb=load_workbook(out_xls); ws=wb.active
        ws.insert_rows(1,6)
        ws["A1"]="Banco: Santander"
        ws["A2"]=f"Empresa: {empresa}"
        ws["A3"]=f"No. Cliente: {no_cli}"
        ws["A4"]=f"Periodo: {periodo}"
        ws["A5"]=f"RFC: {rfc}"

        thin=Side("thin"); border=Border(top=thin,left=thin,right=thin,bottom=thin)
        head_fill=PatternFill("solid","000080"); white=Font(color="FFFFFF",bold=True)

        for c in range(1,ws.max_column+1):
            cell=ws.cell(7,c)
            cell.fill=head_fill; cell.font=white
            cell.alignment=Alignment(horizontal="center"); cell.border=border
        for r in range(8,ws.max_row+1):
            for c in range(1,ws.max_column+1):
                ws.cell(r,c).border=border
        for col in ws.columns:
            width=max(len(str(cell.value)) for cell in col if cell.value)+2
            ws.column_dimensions[col[0].column_letter].width=width
        for row in ws.iter_rows():
            for cell in row: cell.alignment=Alignment(wrap_text=True)

        wb.save(out_xls)
        messagebox.showinfo("Éxito", f"Archivo Excel generado:\n{out_xls}")


# -------------------------- MAIN (igual) ---------------------------
def cargar_archivo():
    global pdf_paths
    archivos=filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=[("Archivos PDF","*.pdf"),("Todos los archivos","*.*")]
    )
    if archivos:
        pdf_paths=list(archivos)
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0,tk.END)
        entry_archivo.insert(0," ; ".join(archivos))
        entry_archivo.config(state=tk.DISABLED)

def procesar_pdf():
    if not pdf_paths:
        messagebox.showwarning("Advertencia","No se ha seleccionado ningún PDF.")
        return
    output_folder = sys.argv[1] if len(sys.argv)>=2 else os.getcwd()
    procesar_archivos(pdf_paths, output_folder)

def main():
    global pdf_paths, entry_archivo
    pdf_paths=[]
    root=tk.Tk()
    root.title("Extracción de Movimientos - Santander")
    root.geometry("600x250")

    tk.Button(root,text="Cargar PDF",command=cargar_archivo,width=30).pack(pady=10)
    entry_archivo=tk.Entry(root,width=80,state=tk.DISABLED)
    entry_archivo.pack(padx=10,pady=10)
    tk.Button(root,text="Procesar PDF",command=procesar_pdf,width=30).pack(pady=10)

    root.mainloop()

if __name__=="__main__":
    main()
