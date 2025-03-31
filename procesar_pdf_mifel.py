import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Variables globales
pdf_paths = ""
entry_archivo = None

##########################################################
# 1) FUNCIONES DE UTILIDAD
##########################################################
REGEX_FECHA = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}$')  # dd/mm/aaaa

def es_linea_movimiento(token):
    """Detecta si un token es una fecha dd/mm/aaaa."""
    return bool(REGEX_FECHA.match(token.strip()))

def es_numero_monetario(texto):
    """Detecta si 'texto' es un número monetario tipo '100,923.30'."""
    return bool(re.match(r'^[\d,]+\.\d{2}$', texto.strip()))

def parse_monetario(txt):
    """Convierte '100,923.30' a float (100923.30)."""
    txt = txt.strip().replace(",", "")
    return float(txt)

def dist(a, b):
    """Calcula la distancia absoluta (para asignar columnas por cercanía)."""
    return abs(a - b)

##########################################################
# 2) INTERFAZ: CARGAR ARCHIVO (Tkinter)
##########################################################
def cargar_archivo():
    global pdf_paths, entry_archivo
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos

##########################################################
# 3) FUNCIÓN: PARSEAR PLANTILLA ANTIGUA
# Formato: Fecha, Referencia, Descripción, RETIROS, DEPÓSITOS, SALDO
##########################################################
def parsear_mifel_antiguo(pdf, pdf_path, output_folder):
    pdf_name = os.path.basename(pdf_path)
    pdf_stem, pdf_ext = os.path.splitext(pdf_name)
    excel_name = pdf_stem + "_MIFEL_ANTIGUO.xlsx"

    if len(pdf.pages) == 0:
        messagebox.showinfo("Info", "El PDF está vacío.")
        return

    # 1) Leer encabezados en la 2da página
    if len(pdf.pages) < 2:
        messagebox.showinfo("Info", "El PDF no tiene segunda página.")
        return

    page1 = pdf.pages[1]
    words_page1 = page1.extract_words()

    col_positions = {}
    headers_needed = ["RETIROS", "DEPÓSITOS", "SALDO"]

    lineas_dict_page1 = {}
    for w in words_page1:
        top_approx = int(w['top'])
        if top_approx not in lineas_dict_page1:
            lineas_dict_page1[top_approx] = []
        lineas_dict_page1[top_approx].append(w)
    lineas_ordenadas_page1 = sorted(lineas_dict_page1.items(), key=lambda x: x[0])
    for top_val, words_in_line in lineas_ordenadas_page1:
        line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
        if all(h in line_text_upper for h in headers_needed):
            for w in words_in_line:
                w_text_up = w['text'].strip().upper()
                if w_text_up in headers_needed:
                    center_x = (w['x0'] + w['x1']) / 2
                    col_positions[w_text_up] = center_x
            break
    columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])
    if not columnas_ordenadas:
        messagebox.showwarning("Advertencia",
            "No se detectaron encabezados en la segunda página. Revisa nombres y acentos.")
        return

    # 2) Extraer datos del encabezado (página 0)
    page0 = pdf.pages[0]
    text_page0_str = page0.extract_text() or ""
    numero_cliente = ""
    match_nc = re.search(r'(?i)Número de cliente\s+(\S+)', text_page0_str)
    if match_nc:
        numero_cliente = match_nc.group(1)
    rfc_cliente = ""
    match_rfc = re.search(r'(?i)RFC\s+([A-Z0-9]{12,13})', text_page0_str)
    if match_rfc:
        rfc_cliente = match_rfc.group(1)
    periodo_str = ""
    match_periodo = re.search(
        r'(?i)DEL\s+(\d{1,2}/\d{1,2}/\d{4})\s+AL\s+(\d{1,2}/\d{1,2}/\d{4})',
        text_page0_str
    )
    if match_periodo:
        periodo_str = match_periodo.group(0)

    # 3) Configuración de frases
    skip_phrases = ["ESTADO DE CUENTA", "PÁGINA", "CARGOS OBJETADOS POR EL CLIENTE",
                    "FONDOS DE INVERSIÓN", "INFORMACIÓN IMPORTANTE", "COMPROBANTE FISCAL", "CUENTA A LA VISTA"]
    skip_phrases = [s.upper() for s in skip_phrases]
    stop_phrases = ["SUMA DE RETIROS Y DEPÓSITOS", "CARGOS OBJETADOS POR EL CLIENTE"]
    stop_phrases = [s.upper() for s in stop_phrases]
    header_phrases = ["FECHA REFERENCIA DESCRIPCIÓN IMPORTE"]
    header_phrases = [s.upper() for s in header_phrases]
    footer_phrases = ["BANCA MIFEL, S.A., INSTITUCIÓN DE BANCA MÚLTIPLE"]
    footer_phrases = [s.upper() for s in footer_phrases]

    start_reading = False
    stop_reading = False
    todos_los_movimientos = []
    movimiento_actual = None

    # 4) Recorrer todas las páginas
    for page_index, page in enumerate(pdf.pages):
        if stop_reading:
            break
        words = page.extract_words()
        lineas_dict = {}
        for w in words:
            top_approx = int(w['top'])
            if top_approx not in lineas_dict:
                lineas_dict[top_approx] = []
            lineas_dict[top_approx].append(w)
        lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])
        for top_val, words_in_line in lineas_ordenadas:
            if stop_reading:
                break
            line_text = " ".join(w['text'] for w in words_in_line)
            line_text_upper = line_text.upper().strip()
            tokens_line = line_text.split()
            if any(fp in line_text_upper for fp in footer_phrases):
                break
            if any(sp in line_text_upper for sp in stop_phrases):
                stop_reading = True
                break
            if not start_reading:
                if len(tokens_line) > 0 and es_linea_movimiento(tokens_line[0]):
                    has_monetary = any(es_numero_monetario(tk) for tk in tokens_line[1:])
                    if has_monetary and not any(sp in line_text_upper for sp in skip_phrases):
                        start_reading = True
                    else:
                        continue
                else:
                    continue
            if any(sp in line_text_upper for sp in skip_phrases):
                continue
            if any(hp in line_text_upper for hp in header_phrases):
                continue
            if "SUMA DE RETIROS" in line_text_upper or "SALDO A FECHA DE CORTE" in line_text_upper:
                continue
            if len(tokens_line) > 0 and es_linea_movimiento(tokens_line[0]):
                has_monetary = any(es_numero_monetario(tk) for tk in tokens_line[1:])
                if not has_monetary:
                    continue
                if movimiento_actual:
                    todos_los_movimientos.append(movimiento_actual)
                movimiento_actual = {
                    "Fecha": tokens_line[0],
                    "Referencia": None,
                    "Descripción": "",
                    "Retiros": None,
                    "Depositos": None,
                    "Saldo": None,
                }
                tokens_line = tokens_line[1:]
            else:
                if not movimiento_actual:
                    movimiento_actual = {
                        "Fecha": "",
                        "Referencia": None,
                        "Descripción": "",
                        "Retiros": None,
                        "Depositos": None,
                        "Saldo": None,
                    }
            for w in words_in_line:
                raw_txt = w['text'].strip()
                if es_numero_monetario(raw_txt):
                    val = parse_monetario(raw_txt)
                    center_w = (w['x0'] + w['x1']) / 2
                    if columnas_ordenadas:
                        col_name, col_xc = min(columnas_ordenadas, key=lambda c: dist(c[1], center_w))
                        if col_name == "RETIROS":
                            movimiento_actual["Retiros"] = val
                        elif col_name == "DEPÓSITOS":
                            movimiento_actual["Depositos"] = val
                        elif col_name == "SALDO":
                            movimiento_actual["Saldo"] = val
                    else:
                        movimiento_actual["Depositos"] = val
            clean_tokens = [t for t in tokens_line if not es_numero_monetario(t)]
            if movimiento_actual["Referencia"] is None and len(clean_tokens) > 0:
                movimiento_actual["Referencia"] = clean_tokens[0]
                desc_tokens = clean_tokens[1:]
            else:
                desc_tokens = clean_tokens
            if desc_tokens:
                movimiento_actual["Descripción"] += " " + " ".join(desc_tokens)
        # fin for lineas_ordenadas
    if movimiento_actual:
        todos_los_movimientos.append(movimiento_actual)
    df = pd.DataFrame(todos_los_movimientos, columns=["Fecha", "Referencia", "Descripción", "Retiros", "Depositos", "Saldo"])
    ruta_salida = os.path.join(output_folder, excel_name)
    df.to_excel(ruta_salida, index=False)
    wb = load_workbook(ruta_salida)
    ws = wb.active
    ws.insert_rows(1, 6)
    ws["A1"] = "Banco: MIFEL - Plantilla Antigua"
    ws["A2"] = f"Núm. Cliente: {numero_cliente}"
    ws["A3"] = f"RFC: {rfc_cliente}"
    ws["A4"] = f"Periodo: {periodo_str}"
    thin_side = Side(border_style="thin")
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    max_row = ws.max_row
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=7, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row in range(8, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(ruta_salida)
    messagebox.showinfo("Éxito", f"Archivo Excel (plantilla antigua) generado: {ruta_salida}")

##########################################################
# 4) FUNCIÓN: PARSEAR PLANTILLA NUEVA
# Formato: Fecha, Descripción, Monto (MXN), Tipo, Folio, Saldo
##########################################################
def parsear_mifel_nuevo(pdf, pdf_path, output_folder):
    pdf_name = os.path.basename(pdf_path)
    pdf_stem, pdf_ext = os.path.splitext(pdf_name)
    excel_name = pdf_stem + "_MIFEL_NUEVO.xlsx"
    if len(pdf.pages) == 0:
        messagebox.showinfo("Info", "El PDF está vacío.")
        return
    headers_needed = ["FECHA", "DESCRIPCIÓN", "MONTO (MXN)", "TIPO", "FOLIO", "SALDO"]
    col_positions = {}
    page0 = pdf.pages[0]
    words_page0 = page0.extract_words()
    lineas_dict_page0 = {}
    for w in words_page0:
        top_approx = int(w['top'])
        if top_approx not in lineas_dict_page0:
            lineas_dict_page0[top_approx] = []
        lineas_dict_page0[top_approx].append(w)
    lineas_ordenadas_page0 = sorted(lineas_dict_page0.items(), key=lambda x: x[0])
    for top_val, words_in_line in lineas_ordenadas_page0:
        line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
        if all(h in line_text_upper for h in headers_needed):
            for w in words_in_line:
                w_text_up = w['text'].strip().upper()
                if w_text_up in headers_needed:
                    center_x = (w['x0'] + w['x1']) / 2
                    col_positions[w_text_up] = center_x
            break
    columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])
    if not columnas_ordenadas:
        messagebox.showwarning("Advertencia",
            "No se detectaron encabezados en la 1ra página para la plantilla nueva.")
        return
    # Procesamos líneas de la plantilla nueva usando tokens y limpiando el signo '$'
    todos_los_movimientos = []
    for page_index, page in enumerate(pdf.pages):
        text = page.extract_text() or ""
        for line in text.splitlines():
            tokens = line.split()
            tokens = [t[1:] if t.startswith("$") else t for t in tokens if t != "$"]
            if len(tokens) < 6:
                continue
            if not es_linea_movimiento(tokens[0]):
                continue
            # Orden asumido:
            # Token 0: Fecha
            # Tokens 1 hasta -4: Descripción
            # Token -4: Monto (MXN)
            # Token -3: Tipo
            # Token -2: Folio
            # Token -1: Saldo
            fecha = tokens[0]
            saldo = tokens[-1]
            folio = tokens[-2]
            monto = tokens[-4]    # Token -4 corresponde al Monto (MXN)
            tipo  = tokens[-3]    # Token -3 corresponde al Tipo
            descripcion = " ".join(tokens[1:-4])
            movimiento = {
                "Fecha": fecha,
                "Descripción": descripcion,
                "Monto (MXN)": monto,
                "Tipo": tipo,
                "Folio": folio,
                "Saldo": saldo
            }
            todos_los_movimientos.append(movimiento)
    df = pd.DataFrame(todos_los_movimientos, columns=["Fecha", "Descripción", "Monto (MXN)", "Tipo", "Folio", "Saldo"])
    ruta_salida = os.path.join(output_folder, excel_name)
    df.to_excel(ruta_salida, index=False)
    wb = load_workbook(ruta_salida)
    ws = wb.active
    ws.insert_rows(1, 6)
    ws["A1"] = "Banco: MIFEL - Plantilla Nueva"
    ws["A2"] = "Información adicional..."
    thin_side = Side(border_style="thin")
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    max_row = ws.max_row
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=7, column=col)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row in range(8, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(ruta_salida)
    messagebox.showinfo("Éxito", f"Archivo Excel (plantilla nueva) generado: {ruta_salida}")

##########################################################
# 5) FUNCIÓN: DETECTAR PLANTILLA Y PROCESAR PDF
##########################################################
def procesar_pdf_main():
    global pdf_paths, output_folder
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return
    for pdf_path in pdf_paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return
                page0 = pdf.pages[0]
                text_page0 = page0.extract_text() or ""
                text_page0_upper = text_page0.upper()
                if "ESTADO DE CUENTA" in text_page0_upper and "CUENTA A LA VISTA" in text_page0_upper:
                    parsear_mifel_antiguo(pdf, pdf_path, output_folder)
                elif "REPORTE DE MOVIMIENTOS DE LA CUENTA TERMINACION" in text_page0_upper:
                    parsear_mifel_nuevo(pdf, pdf_path, output_folder)
                else:
                    messagebox.showwarning("Error", f"No se reconoció la plantilla en el PDF:\n{pdf_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

##########################################################
# 6) MAIN: INTERFAZ TKINTER
##########################################################
def main():
    global pdf_paths, output_folder, entry_archivo
    if len(sys.argv) < 2:
        print("Uso: python procesar_pdf_mifel.py <carpeta_salida>")
        return
    output_folder = sys.argv[1]
    pdf_paths = ""
    root = tk.Tk()
    root.title("Extracción Movimientos - MIFEL (Multi-plantilla)")
    win_width = 600
    win_height = 250
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")
    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))
    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)
    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf_main, width=30)
    btn_procesar.pack(pady=10)
    root.mainloop()

if __name__ == "__main__":
    main()
