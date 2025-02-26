import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
import string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Conjunto de meses cortos
MESES_CORTOS = {"ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
                "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"}

def es_linea_movimiento(linea):
    """
    Determina si la línea inicia un 'movimiento' nuevo
    con el primer token en formato dd-MES-yy (p.ej. '03-ENE-25').
    """
    tokens = linea.split()
    if not tokens:
        return False
    
    primer_token = tokens[0]
    partes = primer_token.split("-")
    if len(partes) != 3:
        return False
    
    dia, mes, anio = partes
    if not re.match(r'^\d{1,2}$', dia):
        return False
    if mes not in MESES_CORTOS:
        return False
    if not re.match(r'^\d{2}$', anio):
        return False

    return True

def es_numero_monetario(texto):
    """
    Determina si un texto es un número tipo '100,923.30'.
    Ajusta si tu PDF usa otro formato.
    """

    pattern = r'^\d{1,3}(,\d{3})*\.\d{2}$'
    return bool(re.match(pattern, texto.strip()))

def dist(a, b):
    """Distancia absoluta entre dos valores."""
    return abs(a - b)

def cargar_archivo():
    global pdf_path
    archivo = filedialog.askopenfilename(
        title="Selecciona un archivo PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivo:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        entry_archivo.insert(0, archivo)
        entry_archivo.config(state=tk.DISABLED)
        pdf_path = archivo

def agrupar_por_top_con_tolerancia(words, tolerancia=2):
    """
    Recibe una lista de 'words' extraídas por pdfplumber y 
    las agrupa en diccionarios de la forma {top_agrupado: [words_en_esa_linea]}.
    La clave 'top_agrupado' es un float o int representativo de la línea.
    
    'tolerancia' indica cuán cerca (en unidades de 'top') deben estar las palabras
    para considerarlas parte de la misma línea.
    """
    lineas_dict = {}

    for w in words:
        actual_top = w['top']
        # Buscamos si hay algún top ya registrado que esté dentro de la tolerancia
        top_encontrado = None
        for top_existente in lineas_dict.keys():
            if abs(top_existente - actual_top) <= tolerancia:
                top_encontrado = top_existente
                break

        if top_encontrado is not None:
            # Agregamos la palabra a la línea existente
            lineas_dict[top_encontrado].append(w)
        else:
            # Creamos una nueva línea
            lineas_dict[actual_top] = [w]

    # Retornamos las líneas ordenadas por el valor de top
    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])
    return lineas_ordenadas

def procesar_pdf():
    global pdf_path
    if not pdf_path:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) == 0:
                messagebox.showinfo("Info", "El PDF está vacío.")
                return

            # 1) DETECTAR ENCABEZADOS EN LA 1RA PÁGINA
            page0 = pdf.pages[0]
            words_page0 = page0.extract_words()

            encabezados_buscar = ["MONTO DEL DEPOSITO", "MONTO DEL RETIRO", "SALDO"]
            col_positions = {}

            # Agrupamos las palabras de la primera página por 'top' para formar líneas
            # lineas_dict_page0 = {}
            # for w in words_page0:
            #     top_approx = int(w['top'])
            #     if top_approx not in lineas_dict_page0:
            #         lineas_dict_page0[top_approx] = []
            #     lineas_dict_page0[top_approx].append(w)

            # lineas_ordenadas_page0 = sorted(lineas_dict_page0.items(), key=lambda x: x[0])

            lineas_ordenadas_page0 = agrupar_por_top_con_tolerancia(words_page0, tolerancia=2)

            # Buscamos la línea que contenga los 3 encabezados
            for top_val, words_in_line in lineas_ordenadas_page0:
                line_text_upper = " ".join(w['text'].strip().upper() for w in words_in_line)
                if all(h in line_text_upper for h in encabezados_buscar):
                    # Extraemos la coordenada de cada encabezado
                    for w in words_in_line:
                        w_text_upper = w['text'].strip().upper()
                        if w_text_upper in encabezados_buscar:
                            center_x = (w['x0'] + w['x1']) / 2
                            col_positions[w_text_upper] = center_x
                    break

            # Ordenamos por la coordenada X
            columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])


            # 2) VARIABLES PARA ENCABEZADOS DEL EXCEL
            periodo_str = ""
            empresa_str = ""
            no_cliente_str = ""
            rfc_str = ""

            # 3) FRASES A OMITIR (skip) Y A DETENER (stop)
            # Probar con las frases en mayusculas
            skip_phrases = [ 
                "ESTADO DE CUENTA",
                "FECHA DE CORTE",
                "Línea Directa para su empresa:",
                "CIUDAD DE MÉXICO: (55)",
                "(81)8156 9640",
                "CIUDAD DE MÉXICO:",
                "Guadalajara",
                "Monterrey",
                "BANCO MERCANTIL DEL NORTE S.A. INSTITUCIÓN DE BANCA MÚLTIPLE GRUPO FINANCIERO BANORTE.",
                "Nuevo Leon. RFC BMN930209927",
                "Resto del país:",
                "DETALLE DE MOVIMIENTOS",
                "Enlace Negocios Basica",
                "Visita nuestra página:",
                "FECHA",
                "DESCRIPCIÓN / ESTABLECIMIENTO",
                "MONTO DEL DEPOSITO",
                "MONTO DEL RETIRO",
                "SALDO",
                "Banco Mercantil del Norte",
            ]
            # skip_phrases = [s.upper() for s in skip_phrases]

            # stop_phrases: solo se detiene si aparece en páginas >= 2
            stop_phrases = ["OTROS"]

            start_reading = False
            stop_reading = False
            todos_los_movimientos = []
            movimiento_actual = None

            # 4) RECORRER TODAS LAS PÁGINAS
            for page_index, page in enumerate(pdf.pages):
                if stop_reading:
                    break

                # if any (sp in line_text_upper for sp in skip_phrases):
                #     continue

                words = page.extract_words()
                lineas_dict = {}
                for w in words:
                    top_approx = int(w['top'])
                    if top_approx not in lineas_dict:
                        lineas_dict[top_approx] = []
                    lineas_dict[top_approx].append(w)

                lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                for top_val, words_in_line in lineas_ordenadas:
                    if stop_reading:
                        break

                    # Construir la línea
                    line_text = " ".join(w['text'] for w in words_in_line)

                    # Insertar espacio entre fecha pegada y texto
                    line_text = re.sub(
                        r'(\d{1,2}-[A-Z]{3}-\d{2})([A-Za-z])',
                        r'\1 \2',
                        line_text
                    )

                    # Convertir a mayúsculas para comparaciones
                    line_text_upper = line_text.upper()

                    # Remover skip_phrases (parcialmente) de la línea
                    # for sp in skip_phrases:
                    #     line_text_upper = line_text_upper.replace(sp, "")

                    # print(f"[DEBUG] Page {page_index}, top {top_val}: {line_text_upper}")

                    # Detectar periodo
                    if "PERIODO" in line_text_upper and not periodo_str:
                        tokens_line = line_text.split()
                        fechas = [t for t in tokens_line if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', t)]
                        if len(fechas) == 2:
                            periodo_str = f"{fechas[0]} al {fechas[1]}"
                        else:
                            periodo_str = line_text
                        continue

                    # Detectar No. de Cliente
                    if "NO. DE CLIENTE:" in line_text_upper and not no_cliente_str:
                        tokens_line = line_text.split()
                        no_cliente_str = tokens_line[-1]
                        continue

                    # Detectar RFC
                    if "RFC:" in line_text_upper and not rfc_str:
                        tokens_line = line_text.split()
                        rfc_str = tokens_line[-1]
                        continue

                    # Stop phrases (solo en páginas >= 2)
                    if page_index >= 1 and any(sp in line_text_upper for sp in stop_phrases):
                        # print("[DEBUG] -> Stop phrase detectada en página >= 2")
                        stop_reading = True
                        break

                    # Empezar a leer movimientos
                    if not start_reading:
                        if es_linea_movimiento(line_text_upper):
                            # print("[DEBUG] -> Se detecta primer movimiento, start_reading = True")
                            start_reading = True
                        else:
                            continue
                    # regresar _upper luego
                    if any(sp in line_text for sp in skip_phrases):
                        # print("[DEBUG] -> Skip phrase detectada, se omite la línea")
                        continue    

                    if re.search(r'\b\d+/\d+\b', line_text_upper):
                        continue

                    # ¿Es nuevo movimiento?
                    if es_linea_movimiento(line_text_upper):
                        # Antes de guardar, limpiar la descripción de skip phrases
                        if movimiento_actual:
                            # for sp in skip_phrases:
                            #     movimiento_actual["Descripción / Establecimiento"] = re.sub(
                            #         re.escape(sp), "", movimiento_actual["Descripción / Establecimiento"], flags=re.IGNORECASE
                            #     )
                            todos_los_movimientos.append(movimiento_actual)

                        tokens_line = line_text_upper.split()
                        movimiento_actual = {
                            "Fecha": tokens_line[0],
                            "Descripción / Establecimiento": "",
                            "Monto del deposito": None,
                            "Monto del retiro": None,
                            "Saldo": None
                        }
                    else:
                        # Continuación de un movimiento
                        if not movimiento_actual:
                            movimiento_actual = {
                                "Fecha": None,
                                "Descripción / Establecimiento": "",
                                "Monto del deposito": None,
                                "Monto del retiro": None,
                                "Saldo": None
                            }

                    # Procesar cada token de la línea
                    for w in words_in_line:

                        token_upper = w['text'].upper()
                        # Si el token contiene alguna de las skip phrases, detener el procesamiento de esta línea
                        if any(sp in token_upper for sp in skip_phrases):
                            break

                        txt = w['text'].strip()
                        center_w = (w['x0'] + w['x1']) / 2

                        if es_numero_monetario(txt):
                            if columnas_ordenadas:
                                col_name, col_center = min(
                                    columnas_ordenadas,
                                    key=lambda x: dist(x[1], center_w)
                                )
                                print(f"[DEBUG] -> Token '{txt}' en columna '{col_name}'")
                                if col_name in "MONTO DEL DEPOSITO":
                                    movimiento_actual["Monto del deposito"] = txt
                                elif col_name in "MONTO DEL RETIRO":
                                    movimiento_actual["Monto del retiro"] = txt
                                elif col_name in "SALDO":
                                    movimiento_actual["Saldo"] = txt
                            else:
                                movimiento_actual["Monto del retiro"] = txt
                        else:
                            # Si el token inicia con un formato fecha, separamos la parte de fecha y el resto
                            m = re.match(r'^(\d{1,2}-[A-Z]{3}-\d{2})(.*)$', txt)
                            if m:
                                date_part = m.group(1)
                                rest = m.group(2)
                                if movimiento_actual["Fecha"] and date_part.upper() == movimiento_actual["Fecha"]:
                                    txt = rest.strip()  # usar solo el remanente
                            clean_txt = txt.strip(string.punctuation)
                            # Omitir tokens que sean solo dígitos, meses o fecha completa
                            if re.match(r'^\d{1,2}$', clean_txt):
                                continue
                            if clean_txt in MESES_CORTOS:
                                continue
                            if re.match(r'^\d{1,2}-[A-Z]{3}-\d{2}$', clean_txt):
                                continue
                            movimiento_actual["Descripción / Establecimiento"] += txt + " "

            # Al terminar, guardar el último movimiento (con limpieza de skip phrases)
            if movimiento_actual:
                for sp in skip_phrases:
                    movimiento_actual["Descripción / Establecimiento"] = re.sub(
                        re.escape(sp), "", movimiento_actual["Descripción / Establecimiento"], flags=re.IGNORECASE
                    )
                todos_los_movimientos.append(movimiento_actual)

        # 5) GUARDAR EN EXCEL
        df = pd.DataFrame(todos_los_movimientos, columns=[
            "Fecha",
            "Descripción / Establecimiento",
            "Monto del deposito",
            "Monto del retiro",
            "Saldo"
        ])

        ruta_salida = "movimientos_banorte_dd_mmm.xlsx"
        df.to_excel(ruta_salida, index=False)

        # Ajustes de estilo con openpyxl
        wb = load_workbook(ruta_salida)
        ws = wb.active

        # Insertar filas para encabezado
        ws.insert_rows(1, 6)

        # Encabezado
        ws["A1"] = f"Banco: Banorte"
        ws["A2"] = f"Empresa: {empresa_str}"
        ws["A3"] = f"No. Cliente: {no_cliente_str}"
        ws["A4"] = f"Periodo: {periodo_str}"
        ws["A5"] = f"RFC: {rfc_str}"

        thin_side = Side(border_style="thin")
        thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)

        max_row = ws.max_row
        max_col = ws.max_column

        # Estilo para la fila de encabezados (fila 7)
        for col in range(1, max_col + 1):
            cell = ws.cell(row=7, column=col)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Estilo para filas de datos
        for row in range(8, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = max_length + 2

        # Alineación wrap_text
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        wb.save(ruta_salida)
        messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")

# Interfaz gráfica con tkinter
root = tk.Tk()
root.title("Extracción Movimientos - dd-MMM-yy con Ajustes")
root.geometry("600x250")

pdf_path = ""

btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
btn_cargar.pack(pady=10)

entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
entry_archivo.pack(padx=10, pady=10)

btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
btn_procesar.pack(pady=10)

root.mainloop()
