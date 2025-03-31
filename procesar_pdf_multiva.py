import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def normalize_minus_signs(text):
    """
    Reemplaza caracteres Unicode que lucen como signo menos
    (– — − etc.) por el ASCII '-' (0x2D).
    """
    # Lista de posibles caracteres "menos" o guiones largos
    minus_variants = ['\u2010', '\u2011', '\u2012', '\u2013', '\u2014', '\u2212']
    for ch in minus_variants:
        text = text.replace(ch, '-')
    return text

def es_fecha_valida_multiva(texto):
    return texto.strip().isdigit() and 1 <= len(texto.strip()) <= 2

def es_linea_movimiento_multiva(linea):
    tokens = linea.split()
    if len(tokens) < 2:
        return False
    if not es_fecha_valida_multiva(tokens[0]):
        return False
    # Asumimos que el identificador de referencia tiene al menos 10 caracteres
    if len(tokens[1]) < 10:
        return False
    return True

def es_numero_monetario(texto):
    """
    Permite números monetarios con signo negativo opcional, separadores de miles (coma)
    y punto decimal con dos dígitos.
    Ej.: -5,000,000.00 o 80,080.75
    """
    return bool(re.match(r'^-?[\d,]+\.\d{2}$', texto.strip()))

def parse_monetario(txt):
    """
    Convierte un texto en un número monetario.
    """
    txt = txt.strip()
    sign = 1

    # Verifica si está entre paréntesis (se asume negativo)
    if txt.startswith("(") and txt.endswith(")"):
        sign = -1
        txt = txt[1:-1].strip()
    # Verifica si empieza con signo "-"
    elif txt.startswith("-"):
        sign = -1
        txt = txt[1:].strip()

    # Elimina comas antes de convertir a float
    txt = txt.replace(",", "")
    return sign * float(txt)

def dist(a, b):
    return abs(a - b)

def build_description(tokens):
    """
    Construye la descripción omitiendo tokens que sean montos monetarios
    o que sean únicamente '-' o '+'.
    """
    desc_tokens = []
    for t in tokens:
        if es_numero_monetario(t):
            continue
        if t.strip() in ["-", "+"]:
            continue
        desc_tokens.append(t)
    return " ".join(desc_tokens)

def unir_signo_con_numero(words_in_line):
    """
    1) Normaliza todos los guiones Unicode a ASCII '-'
    2) Une tokens separados que sean un signo '-' o '+' seguido inmediatamente
       de un token que coincida con un número (ej. 5,000.00), para formar
       '-5,000.00'.
    Retorna una lista de tokens con la posible unificación.
    """
    new_line = []
    skip_next = False

    for i in range(len(words_in_line)):
        if skip_next:
            skip_next = False
            continue

        w = words_in_line[i]
        # Normalizar los guiones en el token actual
        w['text'] = normalize_minus_signs(w['text'])
        txt = w['text'].strip()

        # Revisar el siguiente token
        if i + 1 < len(words_in_line):
            w_next = words_in_line[i + 1]
            w_next['text'] = normalize_minus_signs(w_next['text'])
            txt_next = w_next['text'].strip()

            # Caso: token actual es '-' o '+', y el siguiente es un número
            if txt in ["-", "+"] and re.match(r'^[\d,]+\.\d{2}$', txt_next):
                # Unirlos
                new_txt = txt + txt_next  # Ej.: "-5,000,000.00"
                new_word = {
                    'x0': w['x0'],
                    'x1': w_next['x1'],
                    'top': w['top'],
                    'bottom': w['bottom'],
                    'text': new_txt
                }
                new_line.append(new_word)
                skip_next = True
                continue

        # Si no se unió, agregar tal cual
        new_line.append(w)

    return new_line

def cargar_archivo():
    global pdf_paths
    archivos = filedialog.askopenfilenames(
        title="Selecciona uno o más archivos PDF",
        filetypes=(("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*"))
    )
    if archivos:
        entry_archivo.config(state=tk.NORMAL)
        entry_archivo.delete(0, tk.END)
        files_text = " ; ".join(archivos)
        entry_archivo.insert(0, files_text)
        entry_archivo.config(state=tk.DISABLED)
        pdf_paths = archivos

def procesar_pdf():
    global pdf_paths, output_folder
    if not pdf_paths:
        messagebox.showwarning("Advertencia", "No se ha seleccionado un archivo PDF.")
        return
    
    for pdf_path in pdf_paths:
        try:
            with pdfplumber.open(pdf_path) as pdf:

                pdf_name = os.path.basename(pdf_path)
                pdf_stem, pdf_ext = os.path.splitext(pdf_name)
                excel_name = pdf_stem + ".xlsx"

                if len(pdf.pages) == 0:
                    messagebox.showinfo("Info", "El PDF está vacío.")
                    return

                periodo_str = ""
                num_cuenta_str = ""
                num_cliente_str = ""
                rfc_str = ""
                banco_str = ""
                empresa_str = ""

                # 1) Detectar encabezados en la primera página
                page0_words = pdf.pages[0].extract_words()
                col_positions = {}
                encabezados_buscar = ["RETIROS", "DEPÓSITOS", "SALDO"]

                for w in page0_words:
                    txt_upper = w['text'].strip().upper()
                    center_x = (w['x0'] + w['x1']) / 2
                    if txt_upper in encabezados_buscar:
                        col_positions[txt_upper] = center_x

                columnas_ordenadas = sorted(col_positions.items(), key=lambda x: x[1])

                # 2) skip_phrases, stop_phrases
                skip_phrases = [
                    "REPRESENTACIÓN IMPRESA DE UN CFDI",
                    "COMPROBANTE FISCAL",
                    "AVE",            
                    "COL.",
                    "C.P:",
                    "REGIMEN FISCAL",
                    "USO DE CFDI",
                    "FECHA DE EXPEDICIÓN",
                    "PERIODO DEL",
                    "NÚMERO DE CLIENTE",
                    "RFC",
                    "CUENTA MULTIEMPRESARIAL",
                    "NÚMERO DE CUENTA",
                    "SUCURSAL",
                    "TIPO DE MONEDA",
                    "CLABE",
                    "RESUMEN DE MOVIMIENTOS",
                    "SALDO ANTERIOR",
                    "RETIROS/DEPÓSITOS",
                    "COMISIONES COBRADAS/BONIFICACIONES",
                    "GASTOS DE ADMINISTRACIÓN/BONIFICACIONES",
                    "INTERESES COBRADOS/BONIFICACIONES",
                    "I.V.A. DE TRANSACCIONES/BONIFICACIONES",
                    "I.S.R. RETENIDO/BONIFICACIONES",
                    "SALDO AL CORTE",
                    "SALDO PROMEDIO",
                    "DÍAS TRANSCURRIDOS",
                    "TASA PROMEDIO BRUTA ANUAL",
                    "DÍA REFERENCIA DESCRIPCIÓN",
                    "BANCO MULTIVA SOCIEDAD ANONIMA",
                    "INSTITUCIÓN DE BANCA MÚLTIPLE",
                    "GRUPO FINANCIERO MULTIVA",
                    "CERRADA DE TECAMACHALCO",
                    "TEL",
                    "PÁGINA"
                ]
                skip_phrases = [s.upper() for s in skip_phrases]

                stop_phrases = ["CARGOS OBJETADOS"]

                start_reading = False
                stop_reading = False

                todos_los_movimientos = []
                movimiento_actual = None

                # 3) Recorrer páginas
                for page_index, page in enumerate(pdf.pages):
                    if stop_reading:
                        break

                    words = page.extract_words()
                    # Agrupar por línea
                    lineas_dict = {}
                    for w in words:
                        top_approx = int(w['top'])
                        if top_approx not in lineas_dict:
                            lineas_dict[top_approx] = []
                        lineas_dict[top_approx].append(w)

                    lineas_ordenadas = sorted(lineas_dict.items(), key=lambda x: x[0])

                    for top_val, words_in_line in lineas_ordenadas:
                        if stop_reading:
                            break

                        # Normalizar/Unir tokens
                        words_in_line = unir_signo_con_numero(words_in_line)
                        line_text = " ".join(w['text'] for w in words_in_line)
                        line_text_upper = line_text.upper()

                        # Extraer período
                        if "PERIODO" in line_text_upper and "AL" in line_text_upper:
                            tokens = line_text.split()
                            fechas = [t for t in tokens if re.match(r'^\d{1,2}-[A-Za-z]{3}-\d{4}$', t)]
                            if len(fechas) == 2:
                                periodo_str = f"{fechas[0]} al {fechas[1]}"
                            else:
                                periodo_str = line_text
                            continue

                        # Extraer Número de Cuenta
                        if "NÚMERO DE CUENTA" in line_text_upper and not num_cuenta_str:
                            tokens = line_text.split()
                            num_cuenta_str = tokens[-1]
                            continue

                        # Extraer Número de Cliente
                        if "NÚMERO DE CLIENTE" in line_text_upper and not num_cliente_str:
                            tokens = line_text.split()
                            num_cliente_str = tokens[-1]
                            continue

                        # Extraer RFC
                        if "RFC" in line_text_upper and not rfc_str:
                            tokens = line_text.split()
                            try:
                                idx = tokens.index("RFC") + 1
                                rfc_str = tokens[idx]
                            except:
                                rfc_str = tokens[-1]
                            continue

                        # Extraer Banco
                        if "BANCO MULTIVA SOCIEDAD ANONIMA" in line_text_upper and not banco_str:
                            banco_str = line_text.strip()
                            continue

                        # Extraer Empresa
                        if not empresa_str and top_val < 100 and not any(sp in line_text_upper for sp in skip_phrases):
                            empresa_str = line_text.strip()
                            skip_phrases.append(empresa_str.upper())
                            continue

                        # Revisar stop
                        if any(sp in line_text_upper for sp in stop_phrases):
                            stop_reading = True
                            break

                        # Comenzar a leer movimientos
                        if not start_reading:
                            tokens_line = line_text.split()
                            if tokens_line and tokens_line[0].isdigit():
                                start_reading = True
                            else:
                                continue

                        # skip_phrases
                        if any(sp in line_text_upper for sp in skip_phrases):
                            continue

                        # Detectar inicio de movimiento
                        if es_linea_movimiento_multiva(line_text):
                            if movimiento_actual:
                                todos_los_movimientos.append(movimiento_actual)
                            tokens_line = line_text.split()
                            day = tokens_line[0]
                            reference = tokens_line[1]
                            desc_line = build_description(tokens_line[2:])

                            movimiento_actual = {
                                "Día": day,
                                "Referencia": reference,
                                "Descripción": desc_line,
                                "Retiros": None,
                                "Depósitos": None,
                                "Saldo": None
                            }
                        else:
                            # Continuación
                            if not movimiento_actual:
                                movimiento_actual = {
                                    "Día": None,
                                    "Referencia": None,
                                    "Descripción": "",
                                    "Retiros": None,
                                    "Depósitos": None,
                                    "Saldo": None
                                }
                            tokens_line = line_text.split()
                            desc_line = build_description(tokens_line)
                            movimiento_actual["Descripción"] += " " + desc_line

                        # Asignar montos
                        for w in words_in_line:
                            txt = w['text'].strip()
                            center_w = (w['x0'] + w['x1']) / 2
                            if es_numero_monetario(txt):
                                val = parse_monetario(txt)
                                if columnas_ordenadas:
                                    col_name, col_center = min(
                                        columnas_ordenadas,
                                        key=lambda x: dist(x[1], center_w)
                                    )
                                    if col_name == "RETIROS":
                                        if txt.startswith("-"):
                                            movimiento_actual["Retiros"] = val
                                    elif col_name == "DEPÓSITOS":
                                        movimiento_actual["Depósitos"] = val
                                    elif col_name == "SALDO":
                                        movimiento_actual["Saldo"] = val
                                else:
                                    movimiento_actual["Retiros"] = val

                if movimiento_actual:
                    todos_los_movimientos.append(movimiento_actual)

            # DataFrame
            df = pd.DataFrame(todos_los_movimientos, columns=[
                "Día",
                "Referencia",
                "Descripción",
                "Retiros",
                "Depósitos",
                "Saldo"
            ])

            ruta_salida = os.path.join(output_folder, excel_name)
            df.to_excel(ruta_salida, index=False)

            # Ajustes openpyxl
            wb = load_workbook(ruta_salida)
            ws = wb.active

            # Insertar filas para encabezado
            ws.insert_rows(1, 6)
            ws["A1"] = f"Banco: {banco_str if banco_str else 'Banco Multiva Sociedad Anónima'}"
            ws["A2"] = f"Empresa: {empresa_str}"
            ws["A3"] = f"Número de Cuenta: {num_cuenta_str}"
            ws["A4"] = f"Número de Cliente: {num_cliente_str}"
            ws["A5"] = f"Periodo: {periodo_str}"
            ws["A6"] = f"RFC: {rfc_str}"

            thin_side = Side(border_style="thin")
            thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

            header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)

            max_row = ws.max_row
            max_col = ws.max_column

            # Encabezados
            for col in range(1, max_col + 1):
                cell = ws.cell(row=7, column=col)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Filas de datos
            for row in range(8, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

            # Ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                ws.column_dimensions[col_letter].width = max_length + 2

            # wrap_text
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            wb.save(ruta_salida)
            messagebox.showinfo("Éxito", f"Archivo Excel generado: {ruta_salida}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar el PDF:\n{e}")
def main():
    global pdf_paths, output_folder
    if len(sys.argv) < 2:
        print("Uso: python procesar_pdf_multiva.py <carpeta_salida>")
        return
    output_folder = sys.argv[1]

    pdf_paths = ""
    # Interfaz
    root = tk.Tk()
    root.title("Extracción Movimientos - Banco Multiva")
    win_width = 600
    win_height = 250

    # Obtenemos dimensiones de la pantalla
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

        # Calculamos coordenadas x e y
    x = (screen_width - win_width) // 2
    y = (screen_height - win_height) // 2
    # Ajustamos la geometría: ancho x alto + x + y
    root.geometry(f"{win_width}x{win_height}+{x}+{y}")

    root.update()
    root.lift()
    root.focus_force()
    root.attributes("-topmost", True)
    root.after(10, lambda: root.attributes("-topmost", False))


    btn_cargar = tk.Button(root, text="Cargar PDF", command=cargar_archivo, width=30)
    btn_cargar.pack(pady=10)

    global entry_archivo
    entry_archivo = tk.Entry(root, width=80, state=tk.DISABLED)
    entry_archivo.pack(padx=10, pady=10)

    btn_procesar = tk.Button(root, text="Procesar PDF", command=procesar_pdf, width=30)
    btn_procesar.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main()